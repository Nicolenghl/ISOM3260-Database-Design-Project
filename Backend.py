# -- Install and setup Oracle Instant Client --
! wget -O oracleinstantclient.zip https://download.oracle.com/otn_software/linux/instantclient/19600/instantclient-basic-linux.x64-19.6.0.0.0dbru.zip
! unzip -oq oracleinstantclient.zip
! mkdir -p /opt/oracle
! mv instantclient_19_6 /opt/oracle/        #mv = move
! sh -c "echo /opt/oracle/instantclient_19_6 > /etc/ld.so.conf.d/oracle-instantclient.conf"
! ldconfig

# -- Include cx_Oracle package --
! pip install cx_Oracle

# Import cx_Oracle package and connect to Oracle Database
import cx_Oracle

HOST_NAME = "imz409.ust.hk"
PORT_NUMBER = "1521"
SERVICE_NAME = "imz409"
USERNAME = "LA205"
PASSWORD = "PA914"

dsn_tns = cx_Oracle.makedsn(HOST_NAME, PORT_NUMBER, service_name=SERVICE_NAME)
conn = cx_Oracle.connect(user=USERNAME, password=PASSWORD, dsn=dsn_tns)

c = conn.cursor()   #connect to the database

# Import required libraries for Excel export functionality
import pandas as pd
import io
import base64
from IPython.display import HTML as DisplayHTML
from google.colab.widgets import TabBar
import ipywidgets as widgets
import time
from datetime import date
from ipywidgets import interact, Text, DatePicker, IntText, Label, Layout, Button, Box, GridspecLayout, Checkbox, Output, Password, Dropdown, HBox, VBox, Dropdown, IntText, Output, HTML
from IPython.display import display, clear_output

# Define global CSS and design guidelines for consistent styling
# These variables will be used throughout the application
# Color scheme
PRIMARY_COLOR = '#4CAF50'    # Green for primary actions
SECONDARY_COLOR = '#2196F3'  # Blue for secondary actions
NEUTRAL_COLOR = '#f0f0f0'    # Light gray for neutral actions
HEADER_STYLE = {'font_weight': 'bold', 'font_size': '18px', 'margin-bottom': '10px'}

# Function to create standard section headers
def create_header(text):
    styled_label = HTML(f"<div style='font-weight: bold; font-size: 17px; margin-bottom: 1px;'>{text}</div>")
    return styled_label

# Function to create standard section subheaders
SUBHEADER_STYLE = {'font_weight': 'bold', 'font_size': '13px', 'margin': '1px 0'}
def create_subheader(text):
    styled_label = HTML(f"""
    <div style='font-weight: {SUBHEADER_STYLE['font_weight']};
                 font-size: {SUBHEADER_STYLE['font_size']};
                 margin: {SUBHEADER_STYLE['margin']};'>
        {text}
    </div>
    """)
    return styled_label

# Function to create standard labels
def create_styled_label(text):
    return HTML(f"<div style='font-size: 13px;'>{text}</div>")

# Button layouts and styling
ACTION_BUTTON_LAYOUT = widgets.Layout(width='150px', margin='5px')
MENU_BUTTON_LAYOUT = widgets.Layout(width='250px', margin='5px')
SEARCH_BUTTON_LAYOUT = widgets.Layout(width='130px', margin='5px')

# Function to style buttons consistently
def style_button(button, color, is_bold=True):
    button.style.button_color = color
    if is_bold:
        button.style.font_weight = 'bold'
    return button

# Input field layouts
INPUT_LAYOUT = widgets.Layout(width='300px')
DROPDOWN_LAYOUT = widgets.Layout(width='300px')

# Section spacing
SECTION_SPACING = widgets.HTML("<br>")
SEPARATOR = widgets.HTML("<hr>")

# Define styles
STATUS_MSG_STYLE = {'font_weight': 'bold'}
ERROR_STYLE = {'color': 'red', 'font_weight': 'bold'}
SUCCESS_STYLE = {'color': 'green', 'font_weight': 'bold'}

def create_styled_message(text, status):
    if status == 'error':
        style = f"color: {ERROR_STYLE['color']}; font-weight: {ERROR_STYLE['font_weight']};"
    elif status == 'success':
        style = f"color: {SUCCESS_STYLE['color']}; font-weight: {SUCCESS_STYLE['font_weight']};"
    else:
        style = f"font-weight: {STATUS_MSG_STYLE['font_weight']};"

    return HTML(f"<div style='{style}'>{text}</div>")

# Assume connection to the database is already established
# conn = cx_Oracle.connect(username, password, dsn)
# c = conn.cursor()

# Create the TabBar with all required tabs
tb_all = TabBar([
    'Login', 'Main Menu', 'Manager Dashboard', 'View All Orders',
    'Sales Revenue by Date', 'Sales Quantity by Date', 'Trade-in Discounts by Date',
    'Old Devices Received', 'Out-of-Stock Products', 'Top 10 Customers',
    'Admin - Add Product', 'Admin - Add Officer', 'Admin - View All Officers',
    'Admin - Manage Officer Status', 'Admin - Edit Products', 'Admin - View All Orders','Admin - Manage Orders',
    'Officer - Change Password', 'Officer - View Trade-in Items', 'Officer - Process Order'
])

# Current user (for role-based access)
global current_user
current_user = {'role': 'Administrator'}  # Simplified for example; actual implementation uses login

# A simple function for jumping between tabs
def showPage(tabName):
  with tb_all.output_to(tabName):
    display()

# Function to toggle password visibility
def toggle_password_visibility(checkbox, password_field, plain_text_field):
  if checkbox.value == True:  # Checkbox is checked
    plain_text_field.value = password_field.value
    password_field.layout.display = 'none'
    plain_text_field.layout.display = ''
  else:  # Checkbox is unchecked
    password_field.value = plain_text_field.value
    plain_text_field.layout.display = 'none'
    password_field.layout.display = ''

# Main function to handle checkbox changes
def on_show_password_checkbox_change(change):
    # Identify which checkbox triggered the change
    checkbox_index = checkbox.index(change['owner'])
    toggle_password_visibility(checkbox[checkbox_index], password_fields[checkbox_index][0], password_fields[checkbox_index][1])

# Helper function to clear password fields and checkboxes
def clear_fields(*fields):
    for field in fields:
        if isinstance(field, Checkbox):
            field.value = False  # Reset checkbox
        else:
            field.value = ""  # Clear text fields

##########################################################
################ Login Tab - Begin ######################
##########################################################
with tb_all.output_to('Login'):

  # Create header
  intro_label = create_header("Welcome to Vincent Mobile!")

  # Create input fields with better visibility
  Login_username_label = create_styled_label("Username:")
  Login_username_textbox = Text(
    value='',
    placeholder='Enter your username',
    layout=widgets.Layout(width='250px', padding='5px', margin='5px')
  )

  Login_password_label = create_styled_label("Password:")
  Login_password_textbox = Password(
    value='',
    placeholder='Enter your password',
    layout=widgets.Layout(width='250px', padding='5px', margin='5px')
  )

  plain_text_password = Text(
    value='',
    placeholder='Enter your password',
    layout=widgets.Layout(width='250px', padding='5px', margin='5px')
  )

  # Initially hide the plain text password
  plain_text_password.layout.display = 'none'

  # Status message for login errors
  login_status_label = Label("")
  login_status_label.style = ERROR_STYLE

  # Create login button
  btnLogin = widgets.Button(description="Login", layout=ACTION_BUTTON_LAYOUT)
  style_button(btnLogin, PRIMARY_COLOR)

  # Checkbox to toggle password visibility
  cb1 = Checkbox(description='Show Password', value=False, disabled=False, indent=False)
  checkbox = [cb1]
  password_fields = [(Login_password_textbox,plain_text_password)]

  cb1.observe(on_show_password_checkbox_change, names='value')

  # Display the objects with a clearer layout
  display(widgets.VBox([
    widgets.Box([intro_label], layout=widgets.Layout(display='flex', justify_content='center', margin='20px 0')),
    SECTION_SPACING,

    # Username field with label
    widgets.VBox([
      Login_username_label,
      Login_username_textbox
    ], layout=widgets.Layout(margin='10px 0')),

    # Password field with label and visibility toggle
    widgets.VBox([
      Login_password_label,
      widgets.HBox([Login_password_textbox, plain_text_password]),
      widgets.Box([cb1], layout=widgets.Layout(margin='5px 0'))
    ], layout=widgets.Layout(margin='10px 0')),

    SECTION_SPACING,

    # Login button centered
    widgets.Box([btnLogin], layout=widgets.Layout(display='flex', justify_content='center', margin='20px 0')),

    # Error message area
    widgets.Box([login_status_label], layout=widgets.Layout(margin='10px 0', min_height='20px'))

  ], layout=widgets.Layout(
    padding='30px',
    width='400px',
    margin='0 auto',
    border='1px solid #ddd',
    border_radius='5px'
  )))

  # Define the login handler
  def on_btnLogin_button_clicked(b):
    # Get the current password value based on which field is visible
    password_value = plain_text_password.value if cb1.value else Login_password_textbox.value

    # Check credentials against database
    try:
      c.execute("""
        SELECT * FROM STAFF
        WHERE S_USERNAME = :username AND S_PW = :password
      """, {'username': Login_username_textbox.value, 'password': password_value})

      result = c.fetchone()

      if result:
        if result[8] == 'Inactive':
          login_status_label.value = "This account is inactive. Please contact administrator."
          login_status_label.style = ERROR_STYLE
          return

        # Store the user role in a global variable for use in main menu
        global current_user
        current_user = {
          'id': result[0],
          'username': result[1],
          'email': result[6],  # S_EMAIL is the 7th column
          'role': 'Administrator' if result[7].startswith('A') else 'Officer' # S_ROLE starting with 'A' are admins
        }

        # Clear status and navigate to main menu
        login_status_label.value = ""
        update_menu_by_role() # Update the menu based on user role
        showPage("Main Menu")
      else:
        login_status_label.value = "Invalid username or password. Please try again."
        login_status_label.style = ERROR_STYLE
    except Exception as e:
      login_status_label.value = f"Login error: {str(e)}"
      login_status_label.style = ERROR_STYLE

  btnLogin.on_click(on_btnLogin_button_clicked)
################ Login Tab - End ################

#######################################################
################ Main Menu Tab - Begin ################
#######################################################
with tb_all.output_to('Main Menu'):
    welcome_label = widgets.Label("Welcome! Please select a function...")

    button_layout = widgets.Layout(width='300px', height='50px')

    btnManagerDashboard = widgets.Button(description="Manager Dashboard", layout=button_layout)
    btnAddProduct = widgets.Button(description="Add New Product", layout=button_layout)
    btnAddOfficer = widgets.Button(description="Add New Officer Account", layout=button_layout)
    btnViewAllOfficers = widgets.Button(description="View All Officers", layout=button_layout)
    btnManageOfficerStatus = widgets.Button(description="Manage Officer Status", layout=button_layout)
    btnEditProducts = widgets.Button(description="View/Edit Product Information", layout=button_layout)
    btnViewOrders = widgets.Button(description="View All Orders", layout=button_layout)
    btnManageOrders = widgets.Button(description="Manage Orders", layout=button_layout)
    btnChangePassword = widgets.Button(description="Change Password", layout=button_layout)
    btnViewTradeInItems = widgets.Button(description="View Trade-in Items", layout=button_layout)
    btnProcessOrder = widgets.Button(description="Process Order", layout=button_layout)
    btnLogout = widgets.Button(description="Logout", layout=button_layout)

    def on_btnManagerDashboard_clicked(b):
        showPage("Manager Dashboard")
    btnManagerDashboard.on_click(on_btnManagerDashboard_clicked)

    def on_btnAddProduct_clicked(b):
        showPage("Admin - Add Product")
    btnAddProduct.on_click(on_btnAddProduct_clicked)

    def on_btnAddOfficer_clicked(b):
        showPage("Admin - Add Officer")
    btnAddOfficer.on_click(on_btnAddOfficer_clicked)

    def on_btnViewAllOfficers_clicked(b):
        showPage("Admin - View All Officers")
    btnViewAllOfficers.on_click(on_btnViewAllOfficers_clicked)

    def on_btnManageOfficerStatus_clicked(b):
        showManageOfficerStatusPage()
    btnManageOfficerStatus.on_click(on_btnManageOfficerStatus_clicked)

    def on_btnEditProducts_clicked(b):
        showPage("Admin - Edit Products")
    btnEditProducts.on_click(on_btnEditProducts_clicked)

    def on_btnViewOrders_clicked(b):
        # First navigate to the All Orders tab
        showPage('Admin - View All Orders')
        # Then explicitly trigger the load orders function after a small delay
        import time
        time.sleep(0.2)  # Short delay to ensure the UI is ready
        # Explicitly call the orders load function - not the officers one
        with tb_all.output_to('Admin - View All Orders'):
            on_VIEW_all_button_clicked(None)
    btnViewOrders.on_click(on_btnViewOrders_clicked)

    def on_btnManageOrders_clicked(b):
        showPage("Admin - Manage Orders")
    btnManageOrders.on_click(on_btnManageOrders_clicked)

    def on_btnChangePassword_clicked(b):
        setup_password_change_tab()
        showPage("Officer - Change Password")
    btnChangePassword.on_click(on_btnChangePassword_clicked)

    def on_btnViewTradeInItems_clicked(b):
        showPage("Officer - View Trade-in Items")
    btnViewTradeInItems.on_click(on_btnViewTradeInItems_clicked)

    def on_btnProcessOrder_clicked(b):
        showPage("Officer - Process Order")
    btnProcessOrder.on_click(on_btnProcessOrder_clicked)

    def on_btnLogout_clicked(b):
        clear_fields(Login_username_textbox, Login_password_textbox, login_status_label, plain_text_password, cb1)
        showPage("Login")
    btnLogout.on_click(on_btnLogout_clicked)

    button_container = widgets.VBox([])
    def update_menu_by_role():
        if current_user['role'] == 'Administrator':
            button_container.children = [btnManagerDashboard, btnAddProduct, btnAddOfficer, btnViewAllOfficers,
                                        btnManageOfficerStatus, btnEditProducts, btnViewOrders, btnManageOrders, btnLogout]
            welcome_label.value = f"Welcome Administrator {current_user['username']}! Please select a function..."
        else:
            button_container.children = [btnChangePassword, btnViewTradeInItems, btnProcessOrder, btnLogout]
            welcome_label.value = f"Welcome Officer {current_user['username']}! Please select a function..."

    display(widgets.VBox([welcome_label, button_container]))

################ Main Menu Tab - End ################

##############################################################
################ Admin - Add Product Tab - Begin #############
##############################################################
with tb_all.output_to('Admin - Add Product'):
  # Get the next available Product ID by finding gaps in existing IDs
  c.execute("SELECT P_ID FROM PRODUCT ORDER BY P_ID")
  existing_ids = [row[0] for row in c.fetchall()]

  next_p_id = 1
  if existing_ids:
      # Find the first gap in the sequence
      for i, p_id in enumerate(existing_ids, start=1):
          if p_id != i:
              next_p_id = i
              break
      else:
          # If no gaps found, use max_id + 1
          next_p_id = max(existing_ids) + 1

  # Create labels and input fields
  PROD_id_label = Label("Product ID: ")
  PROD_id_textbox = Text(value=str(next_p_id), placeholder='Auto-generated', disabled=True)

  PROD_name_label = Label("Product Name: ")
  PROD_name_textbox = Text(value='', placeholder='Enter Product Name', disabled=False)

  PROD_desc_label = Label("Description: ")
  PROD_desc_textbox = Text(value='', placeholder='Enter Product Description', disabled=False)

  PROD_price_label = Label("Price: ")
  PROD_price_textbox = Text(value='', placeholder='Enter Price (e.g., 19.99)', disabled=False)

  # Weight field - users just enter the number
  PROD_weight_label = Label("Weight (grams): ")
  PROD_weight_textbox = Text(value='', placeholder='Enter weight number only', disabled=False)

  PROD_colour_label = Label("Colour: ")
  PROD_colour_textbox = Text(value='', placeholder='Enter Colour', disabled=False)

  # Capacity field with number input and unit dropdown
  PROD_capacity_label = Label("Capacity: ")
  PROD_capacity_textbox = Text(value='', placeholder='Enter number', disabled=False)
  PROD_capacity_dropdown = widgets.Dropdown(options=['GB', 'TB'], value='GB', description='')
  PROD_capacity_box = widgets.HBox([PROD_capacity_textbox, PROD_capacity_dropdown])

  PROD_manufacturer_label = Label("Manufacturer: ")
  PROD_manufacturer_textbox = Text(value='', placeholder='Enter Manufacturer', disabled=False)

  PROD_status_label = Label("Status: ")
  PROD_status_dropdown = widgets.Dropdown(
    options=['Active', 'Inactive'],
    value='Active',
    description='',
  )

  PROD_inventory_label = Label("Inventory: ")
  PROD_inventory_intbox = IntText(value=0, description='', disabled=False)

  # Status message
  PROD_status_msg = Label("")

  # Create buttons
  PROD_add_button = widgets.Button(description="Add Product")
  PROD_clear_button = widgets.Button(description="Clear")
  PROD_back_button = widgets.Button(description="Back to Menu")

  # Define button handlers
  def on_PROD_add_button_clicked(b):
    # Validate all fields are filled
    required_fields = [
        (PROD_name_textbox, "Product Name"),
        (PROD_desc_textbox, "Description"),
        (PROD_price_textbox, "Price"),
        (PROD_weight_textbox, "Weight"),
        (PROD_colour_textbox, "Colour"),
        (PROD_capacity_textbox, "Capacity"),
        (PROD_manufacturer_textbox, "Manufacturer")
    ]

    missing_fields = [name for field, name in required_fields if not field.value.strip()]

    if missing_fields:
        PROD_status_msg.value = f"Please fill in all fields before continuing. Missing: {', '.join(missing_fields)}"
        return

    try:
      # Combine capacity value with unit without space
      capacity_value = f"{PROD_capacity_textbox.value}{PROD_capacity_dropdown.value}"

      # Get weight value and automatically append 'g'
      weight_value = f"{PROD_weight_textbox.value}g"  # Automatically add 'g' here

      # Check inventory and set status accordingly
      if PROD_inventory_intbox.value == 0:
        PROD_status_dropdown.value = "Inactive"
      else: PROD_status_dropdown.value = "Active"

      c.execute(
      """
        INSERT INTO PRODUCT (P_ID, P_NAME, P_DESCRIPTION, P_PRICE, P_WEIGHT, P_COLOUR, P_CAPACITY, P_MANUFACTURER, P_STATUS, P_INVENTORY)
        VALUES ('{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', {})
      """.format(
          PROD_id_textbox.value,
          PROD_name_textbox.value,
          PROD_desc_textbox.value,
          PROD_price_textbox.value,
          weight_value,  # Now using the automatically formatted weight value
          PROD_colour_textbox.value,
          capacity_value,
          PROD_manufacturer_textbox.value,
          PROD_status_dropdown.value,
          PROD_inventory_intbox.value
      ))

      c.execute("commit")
      PROD_status_msg.value = "Product added successfully!"

      # Find the next available ID (checking for gaps again)
      c.execute("SELECT P_ID FROM PRODUCT ORDER BY P_ID")
      existing_ids = [row[0] for row in c.fetchall()]

      next_p_id = 1
      if existing_ids:
          # Find the first gap in the sequence
          for i, p_id in enumerate(existing_ids, start=1):
              if p_id != i:
                  next_p_id = i
                  break
          else:
              # If no gaps found, use max_id + 1
              next_p_id = max(existing_ids) + 1

      PROD_id_textbox.value = str(next_p_id)
      # Clear other fields (except ID)
      on_PROD_clear_button_clicked(None)
    except Exception as e:
      PROD_status_msg.value = f"Error adding product: {str(e)}"

  PROD_add_button.on_click(on_PROD_add_button_clicked)

  def on_PROD_clear_button_clicked(b):
    # Don't clear the auto-generated ID
    PROD_name_textbox.value = ""
    PROD_desc_textbox.value = ""
    PROD_price_textbox.value = ""
    PROD_weight_textbox.value = ""
    PROD_colour_textbox.value = ""
    PROD_capacity_textbox.value = ""
    PROD_capacity_dropdown.value = "GB"
    PROD_manufacturer_textbox.value = ""
    PROD_status_dropdown.value = "Active"
    PROD_inventory_intbox.value = 0
    PROD_status_msg.value = ""

  PROD_clear_button.on_click(on_PROD_clear_button_clicked)

  def on_PROD_back_button_clicked(b):
    showPage("Main Menu")

  PROD_back_button.on_click(on_PROD_back_button_clicked)

  # Display the objects
  PROD_labels = widgets.VBox([PROD_id_label, PROD_name_label, PROD_desc_label, PROD_price_label,
                            PROD_weight_label, PROD_colour_label, PROD_capacity_label,
                            PROD_manufacturer_label, PROD_status_label, PROD_inventory_label])

  PROD_inputs = widgets.VBox([PROD_id_textbox, PROD_name_textbox, PROD_desc_textbox, PROD_price_textbox,
                            PROD_weight_textbox, PROD_colour_textbox, PROD_capacity_box,
                            PROD_manufacturer_textbox, PROD_status_dropdown, PROD_inventory_intbox])

  PROD_form = widgets.HBox([PROD_labels, PROD_inputs])
  PROD_buttons = widgets.HBox([PROD_add_button, PROD_clear_button])

  display(widgets.VBox([PROD_form, PROD_status_msg, PROD_buttons, PROD_back_button]))
################ Admin - Add Product Tab - End ################

##############################################################
################ Admin - Add Officer Tab - Begin #############
##############################################################
with tb_all.output_to('Admin - Add Officer'):
    # Get the next available Officer ID by finding the maximum existing ID
    c.execute("SELECT MAX(S_ID) FROM STAFF")
    max_id = c.fetchone()[0]
    next_officer_id = 1 if max_id is None else max_id + 1  # Start from 1 or increment max ID

    # Create labels and input fields
    OFF_id_label = Label("Officer ID: ")
    OFF_id_textbox = Text(value=str(next_officer_id), placeholder='Auto-generated', disabled=True)

    OFF_username_label = Label("Username: ")
    OFF_username_textbox = Text(value='', placeholder='Enter Username', disabled=False)

    OFF_password_label = Label("Password: ")
    OFF_password_textbox = Text(value='', placeholder='Enter Password', disabled=False)

    OFF_fname_label = Label("First Name: ")
    OFF_fname_textbox = Text(value='', placeholder='Enter First Name', disabled=False)

    OFF_lname_label = Label("Last Name: ")
    OFF_lname_textbox = Text(value='', placeholder='Enter Last Name', disabled=False)

    OFF_phone_label = Label("Phone: ")
    OFF_phone_textbox = Text(value='', placeholder='Enter Phone Number', disabled=False)

    OFF_email_label = Label("Email: ")
    OFF_email_textbox = Text(value='', placeholder='Enter Email', disabled=False)

    OFF_role_label = Label("Role: ")
    OFF_role_textbox = Text(value='Officer', placeholder='Officer', disabled=True)

    OFF_status_label = Label("Status: ")
    OFF_status_dropdown = widgets.Dropdown(
        options=['Active', 'Inactive'],
        value='Active',
        description='',
    )

    # Status message
    OFF_status_msg = Label("")

    # Create buttons
    OFF_add_button = widgets.Button(description="Add Officer")
    OFF_clear_button = widgets.Button(description="Clear")
    OFF_back_button = widgets.Button(description="Back to Menu")

    # Define button handlers
    def on_OFF_add_button_clicked(b):
        # Validate all fields are filled
        required_fields = [
            (OFF_username_textbox, "Username"),
            (OFF_password_textbox, "Password"),
            (OFF_fname_textbox, "First Name"),
            (OFF_lname_textbox, "Last Name"),
            (OFF_phone_textbox, "Phone"),
            (OFF_email_textbox, "Email")
        ]

        missing_fields = [name for field, name in required_fields if not field.value.strip()]

        if missing_fields:
            OFF_status_msg.value = f"Please fill in all fields before continuing. Missing: {', '.join(missing_fields)}"
            return

        try:
            # First check if username already exists
            c.execute("SELECT COUNT(*) FROM STAFF WHERE S_USERNAME = :username",
                     {'username': OFF_username_textbox.value})
            username_exists = c.fetchone()[0]

            if username_exists > 0:
                OFF_status_msg.value = "This username already exists. Please choose another one."
                return

            # Insert new officer into the database
            c.execute("""
                INSERT INTO STAFF (S_ID, S_USERNAME, S_PW, FNAME, LNAME, S_PHONE, S_EMAIL, S_ROLE, S_STATUS)
                VALUES (:id, :username, :password, :fname, :lname, :phone, :email, 'Officer', :status)
            """,
            {
                'id': OFF_id_textbox.value,
                'username': OFF_username_textbox.value,
                'password': OFF_password_textbox.value,
                'fname': OFF_fname_textbox.value,
                'lname': OFF_lname_textbox.value,
                'phone': OFF_phone_textbox.value,
                'email': OFF_email_textbox.value,
                'status': OFF_status_dropdown.value
            })

            c.execute("commit")
            OFF_status_msg.value = "Officer added successfully!"
            # Auto-increment the ID for next entry
            OFF_id_textbox.value = str(int(OFF_id_textbox.value) + 1)
            # Clear other fields (except ID)
            on_OFF_clear_button_clicked(None)
        except Exception as e:
            OFF_status_msg.value = f"Error adding officer: {str(e)}"
            # Print the full error for debugging
            print(f"Full error details: {e}")

    OFF_add_button.on_click(on_OFF_add_button_clicked)

    def on_OFF_clear_button_clicked(b):
        # Don't clear the auto-generated ID
        OFF_username_textbox.value = ""
        OFF_password_textbox.value = ""
        OFF_fname_textbox.value = ""
        OFF_lname_textbox.value = ""
        OFF_phone_textbox.value = ""
        OFF_email_textbox.value = ""
        OFF_status_dropdown.value = "Active"
        OFF_status_msg.value = ""

    OFF_clear_button.on_click(on_OFF_clear_button_clicked)

    def on_OFF_back_button_clicked(b):
        showPage("Main Menu")

    OFF_back_button.on_click(on_OFF_back_button_clicked)

    # Display the objects
    OFF_labels = widgets.VBox([OFF_id_label, OFF_username_label, OFF_password_label, OFF_fname_label,
                             OFF_lname_label, OFF_phone_label, OFF_email_label, OFF_role_label, OFF_status_label])
    OFF_inputs = widgets.VBox([OFF_id_textbox, OFF_username_textbox, OFF_password_textbox, OFF_fname_textbox,
                             OFF_lname_textbox, OFF_phone_textbox, OFF_email_textbox, OFF_role_textbox, OFF_status_dropdown])

    OFF_form = widgets.HBox([OFF_labels, OFF_inputs])
    OFF_buttons = widgets.HBox([OFF_add_button, OFF_clear_button])

    display(widgets.VBox([OFF_form, OFF_status_msg, OFF_buttons, OFF_back_button]))
################ Admin - Add Officer Tab - End ################

##############################################################
########### Admin - View Officers Tab - Begin ################
##############################################################
with tb_all.output_to('Admin - View All Officers'):

  # Create header with title using standard style
  VIEW_title = create_header("View All Officers")

  # Create retrieve all button with standard layout (we'll hide it but keep it for functionality)
  VIEW_all_button = widgets.Button(
    description="Load All Officers",
    layout=widgets.Layout(display='none')  # Hide the button
  )
  VIEW_back_button = widgets.Button(
    description="Back to Menu",
    layout=ACTION_BUTTON_LAYOUT
  )

  # Apply styles to buttons using the helper function
  style_button(VIEW_all_button, PRIMARY_COLOR)
  style_button(VIEW_back_button, NEUTRAL_COLOR)

  # Create a refresh button to manually refresh the data if needed
  VIEW_refresh_button = widgets.Button(
    description="Refresh",
    layout=ACTION_BUTTON_LAYOUT
  )
  style_button(VIEW_refresh_button, SECONDARY_COLOR)

  # Create output for displaying all officers
  VIEW_all_output = widgets.Output()

  # Create message for displaying number of retrieved officers
  VIEW_all_count_msg = widgets.Label("")
  VIEW_all_count_msg.style = STATUS_MSG_STYLE

  # Lists for storing textboxes when displaying all officers
  VIEW_all_id_textboxes = []
  VIEW_all_username_textboxes = []
  VIEW_all_fname_textboxes = []
  VIEW_all_lname_textboxes = []
  VIEW_all_phone_textboxes = []
  VIEW_all_email_textboxes = []
  VIEW_all_role_textboxes = []
  VIEW_all_status_textboxes = []
  VIEW_all_HBoxes = []
  VIEW_all_update_buttons = []  # Define this list globally

  # Status message for errors
  VIEW_status_msg = widgets.Label("")
  VIEW_status_msg.style = ERROR_STYLE

  # Create a flag to track if the View Officers tab needs refreshing
  refresh_view_officers = False

  def on_VIEW_all_button_clicked(b):
    try:
      # Clear previous data
      VIEW_all_id_textboxes.clear()
      VIEW_all_username_textboxes.clear()
      VIEW_all_fname_textboxes.clear()
      VIEW_all_lname_textboxes.clear()
      VIEW_all_phone_textboxes.clear()
      VIEW_all_email_textboxes.clear()
      VIEW_all_role_textboxes.clear()
      VIEW_all_status_textboxes.clear()
      VIEW_all_HBoxes.clear()
      VIEW_all_update_buttons.clear()  # Clear update buttons too

      VIEW_all_output.clear_output()
      VIEW_all_count_msg.value = ""
      VIEW_status_msg.value = ""

      # Debugging: Check if database connection 'c' is available
      if 'c' not in globals():
        try:
          # Try to get connection from the connection pool if available
          global c, conn
          if 'conn' in globals() and conn is not None:
            c = conn.cursor()
          else:
            VIEW_status_msg.value = "Database connection is not available. Please try again."
            return
        except Exception as conn_error:
          VIEW_status_msg.value = f"Database connection error: {str(conn_error)}"
          return

      # Retrieve all staff from database - Order by S_ID ascending
      try:
        c.execute("""
          SELECT * FROM STAFF
          WHERE S_ROLE = 'Officer'
          ORDER BY S_ID ASC
        """)

        rows = c.fetchall()
        if not rows:
          VIEW_all_count_msg.value = "No officers found in the database."
          return
      except Exception as sql_error:
        VIEW_status_msg.value = f"Database query error: {str(sql_error)}"
        return

      # Continue with displaying results...
      with VIEW_all_output:
        from IPython.display import clear_output
        clear_output(wait=True)

        # Define consistent widths for better alignment
        id_width = '60px'
        username_width = '100px'
        fname_width = '100px'
        lname_width = '100px'
        phone_width = '100px'
        email_width = '180px'
        role_width = '80px'
        status_width = '80px'
        button_width = '90px'

        # Apply consistent styles to all cells
        header_style = {'font_weight': 'bold', 'background': '#f2f2f2', 'padding': '5px'}
        cell_style = {'padding': '3px'}

        # Create a header row with same widths as data cells
        header = widgets.HBox([
          widgets.Label("ID", layout=widgets.Layout(width=id_width), style=header_style),
          widgets.Label("Username", layout=widgets.Layout(width=username_width), style=header_style),
          widgets.Label("First Name", layout=widgets.Layout(width=fname_width), style=header_style),
          widgets.Label("Last Name", layout=widgets.Layout(width=lname_width), style=header_style),
          widgets.Label("Phone", layout=widgets.Layout(width=phone_width), style=header_style),
          widgets.Label("Email", layout=widgets.Layout(width=email_width), style=header_style),
          widgets.Label("Role", layout=widgets.Layout(width=role_width), style=header_style),
          widgets.Label("Status", layout=widgets.Layout(width=status_width), style=header_style),
          widgets.Label("Action", layout=widgets.Layout(width=button_width), style=header_style)
        ])

        # Display results with consistent widths
        for row in rows:
          # Apply alternating row colors for better readability
          row_style = {'background': '#f9f9f9'} if len(VIEW_all_id_textboxes) % 2 == 0 else {'background': '#ffffff'}

          # Create textboxes with consistent styling
          id_text = widgets.Text(value=str(row[0]), layout=widgets.Layout(width=id_width, padding='2px'), disabled=True)
          id_text.style = row_style
          VIEW_all_id_textboxes.append(id_text)

          username_text = widgets.Text(value=str(row[1]), layout=widgets.Layout(width=username_width, padding='2px'), disabled=True)
          username_text.style = row_style
          VIEW_all_username_textboxes.append(username_text)

          fname_text = widgets.Text(value=str(row[3]), layout=widgets.Layout(width=fname_width, padding='2px'), disabled=True)
          fname_text.style = row_style
          VIEW_all_fname_textboxes.append(fname_text)

          lname_text = widgets.Text(value=str(row[4]), layout=widgets.Layout(width=lname_width, padding='2px'), disabled=True)
          lname_text.style = row_style
          VIEW_all_lname_textboxes.append(lname_text)

          phone_text = widgets.Text(value=str(row[5]), layout=widgets.Layout(width=phone_width, padding='2px'), disabled=True)
          phone_text.style = row_style
          VIEW_all_phone_textboxes.append(phone_text)

          email_text = widgets.Text(value=str(row[6]), layout=widgets.Layout(width=email_width, padding='2px'), disabled=True)
          email_text.style = row_style
          VIEW_all_email_textboxes.append(email_text)

          role_text = widgets.Text(value=str(row[7]), layout=widgets.Layout(width=role_width, padding='2px'), disabled=True)
          role_text.style = row_style
          VIEW_all_role_textboxes.append(role_text)

          status_text = widgets.Text(value=str(row[8]), layout=widgets.Layout(width=status_width, padding='2px'), disabled=True)
          status_text.style = row_style
          VIEW_all_status_textboxes.append(status_text)

          # Create update button for each row
          update_btn = widgets.Button(
              description="Update",
              button_style='info',
              layout=widgets.Layout(width=button_width, padding='2px')
          )
          update_btn.officer_id = str(row[0])  # Store officer ID as attribute
          update_btn.style.font_weight = 'bold'
          update_btn.on_click(on_row_update_button_clicked)
          VIEW_all_update_buttons.append(update_btn)

        # Create HBoxes for each row with update button
        for i in range(len(VIEW_all_id_textboxes)):
          row_box = widgets.HBox([
            VIEW_all_id_textboxes[i],
            VIEW_all_username_textboxes[i],
            VIEW_all_fname_textboxes[i],
            VIEW_all_lname_textboxes[i],
            VIEW_all_phone_textboxes[i],
            VIEW_all_email_textboxes[i],
            VIEW_all_role_textboxes[i],
            VIEW_all_status_textboxes[i],
            VIEW_all_update_buttons[i]
          ])
          row_box.layout.margin = '0'
          row_box.layout.padding = '0'
          VIEW_all_HBoxes.append(row_box)

        # Add a container for the table with a border
        table_container = widgets.VBox([header] + [i for i in VIEW_all_HBoxes])
        table_container.layout.border = '1px solid #ddd'

        # Display all officers
        from IPython.display import display
        display(table_container)

        # Update count message
        VIEW_all_count_msg.value = f"Retrieved {len(VIEW_all_id_textboxes)} officers."

    except Exception as e:
      import traceback
      VIEW_status_msg.value = f"Error retrieving officers: {str(e)}\n{traceback.format_exc()}"

  # Function to handle row update button clicks
  def on_row_update_button_clicked(b):
    officer_id = b.officer_id
    # Navigate to Manage Officer Status tab
    showPage("Admin - Manage Officer Status")
    # Pre-populate the search field with the officer ID
    MANAGE_search_textbox.value = officer_id
    # Trigger the search button click to load officer details
    on_MANAGE_search_button_clicked(None)

  # Set up button handlers
  VIEW_all_button.on_click(on_VIEW_all_button_clicked)
  VIEW_refresh_button.on_click(on_VIEW_all_button_clicked)  # Use same handler for refresh button

  def on_VIEW_back_button_clicked(b):
    showPage("Main Menu")

  VIEW_back_button.on_click(on_VIEW_back_button_clicked)

  # Modified display for View All Officers tab with consistent spacing
  display(widgets.VBox([
    VIEW_title,
    widgets.HBox([VIEW_refresh_button, VIEW_back_button], layout=widgets.Layout(margin='10px 0')),
    SECTION_SPACING,
    VIEW_all_output,
    VIEW_all_count_msg,
    VIEW_status_msg
  ], layout=widgets.Layout(padding='20px')))

  # Auto-load the officers when the page is opened
  on_VIEW_all_button_clicked(None)
##############################################################
########### Admin - View Officers Tab - End ##################
##############################################################

##############################################################
########### Admin - Manage Officer Status Tab - Begin ########
##############################################################
with tb_all.output_to('Admin - Manage Officer Status'):

  # Header and description
  MANAGE_title = create_header("Update Officer Status")
  MANAGE_description = create_subheader("Search for an officer and update their status")

  # Create search section with better styling
  MANAGE_search_label =create_styled_label("Officer ID: ")
  MANAGE_search_textbox = Text(value='', placeholder='Enter Officer ID', layout=INPUT_LAYOUT, disabled=False)
  MANAGE_search_button = widgets.Button(
    description="Search Officer",
    layout=SEARCH_BUTTON_LAYOUT
  )
  style_button(MANAGE_search_button, PRIMARY_COLOR)

  # Status message for search and update
  MANAGE_status_msg = Label("")
  MANAGE_status_msg.style = STATUS_MSG_STYLE

  # Create display section with separator
  MANAGE_spacer = SEPARATOR

  # Create fields for viewing officer info with consistent styling
  MANAGE_id_label = create_styled_label("Officer ID: ")
  MANAGE_id_textbox = Text(value='', placeholder='Officer ID', layout=INPUT_LAYOUT, disabled=True)

  MANAGE_username_label = create_styled_label("Username: ")
  MANAGE_username_textbox = Text(value='', placeholder='Username', layout=INPUT_LAYOUT, disabled=True)

  MANAGE_fname_label = create_styled_label("First Name: ")
  MANAGE_fname_textbox = Text(value='', placeholder='First Name', layout=INPUT_LAYOUT, disabled=True)

  MANAGE_lname_label = create_styled_label("Last Name: ")
  MANAGE_lname_textbox = Text(value='', placeholder='Last Name', layout=INPUT_LAYOUT, disabled=True)

  MANAGE_phone_label = create_styled_label("Phone: ")
  MANAGE_phone_textbox = Text(value='', placeholder='Phone', layout=INPUT_LAYOUT, disabled=True)

  MANAGE_email_label = create_styled_label("Email: ")
  MANAGE_email_textbox = Text(value='', placeholder='Email', layout=INPUT_LAYOUT, disabled=True)

  MANAGE_role_label = create_styled_label("Role: ")
  MANAGE_role_textbox = Text(value='', placeholder='Role', layout=INPUT_LAYOUT, disabled=True)

  MANAGE_status_label = create_styled_label("Status: ")
  MANAGE_status_dropdown = widgets.Dropdown(
    options=['Active', 'Inactive'],
    value='Active',
    description='',
    layout=DROPDOWN_LAYOUT,
    disabled=False
  )


  # Create buttons with standardized layout and styling
  MANAGE_update_button = widgets.Button(
    description="Update Status",
    disabled=True,
    layout=ACTION_BUTTON_LAYOUT
  )
  MANAGE_clear_button = widgets.Button(
    description="Clear Form",
    layout=ACTION_BUTTON_LAYOUT
  )
  MANAGE_back_button = widgets.Button(
    description="Back to Menu",
    layout=ACTION_BUTTON_LAYOUT
  )
  MANAGE_return_button = widgets.Button(
    description="Return to View",
    layout=ACTION_BUTTON_LAYOUT
  )

  # Apply consistent button styling
  style_button(MANAGE_update_button, PRIMARY_COLOR)
  style_button(MANAGE_clear_button, NEUTRAL_COLOR)
  style_button(MANAGE_back_button, NEUTRAL_COLOR)
  style_button(MANAGE_return_button, SECONDARY_COLOR)

  # Define button handlers
  def on_MANAGE_search_button_clicked(b):
    try:
      # Search by ID
      c.execute("""
        SELECT * FROM STAFF
        WHERE S_ID = :id AND S_ROLE = 'Officer'
      """, {'id': MANAGE_search_textbox.value})

      row = c.fetchone()
      if row:
        MANAGE_id_textbox.value = str(row[0])
        MANAGE_username_textbox.value = row[1]
        MANAGE_fname_textbox.value = row[3]
        MANAGE_lname_textbox.value = row[4]
        MANAGE_phone_textbox.value = row[5]
        MANAGE_email_textbox.value = row[6]
        MANAGE_role_textbox.value = row[7]
        MANAGE_status_dropdown.value = row[8]
        MANAGE_update_button.disabled = False
        MANAGE_status_msg.value = f"Officer found: {row[3]} {row[4]}"
        MANAGE_status_msg.style = SUCCESS_STYLE
      else:
        MANAGE_status_msg.value = "No officer found with that ID. Please check the ID and try again."
        MANAGE_status_msg.style = ERROR_STYLE
        MANAGE_update_button.disabled = True
        clear_officer_fields()
    except Exception as e:
      MANAGE_status_msg.value = f"Error searching for officer: {str(e)}"
      MANAGE_status_msg.style = ERROR_STYLE
      MANAGE_update_button.disabled = True

  MANAGE_search_button.on_click(on_MANAGE_search_button_clicked)

  def on_MANAGE_update_button_clicked(b):
    global refresh_view_officers

    try:
      c.execute(
      """
        UPDATE STAFF
        SET S_STATUS = :status
        WHERE S_ID = :id
      """, {'status': MANAGE_status_dropdown.value, 'id': MANAGE_id_textbox.value})

      c.execute("commit")
      MANAGE_status_msg.value = f"Officer status updated successfully to {MANAGE_status_dropdown.value}!"
      MANAGE_status_msg.style = SUCCESS_STYLE

      # Set the flag to refresh the View Officers tab
      refresh_view_officers = True

    except Exception as e:
      MANAGE_status_msg.value = f"Error updating officer status: {str(e)}"
      MANAGE_status_msg.style = ERROR_STYLE

  MANAGE_update_button.on_click(on_MANAGE_update_button_clicked)

  def clear_officer_fields():
    MANAGE_id_textbox.value = ""
    MANAGE_username_textbox.value = ""
    MANAGE_fname_textbox.value = ""
    MANAGE_lname_textbox.value = ""
    MANAGE_phone_textbox.value = ""
    MANAGE_email_textbox.value = ""
    MANAGE_role_textbox.value = ""
    MANAGE_status_dropdown.value = "Active"

  def on_MANAGE_clear_button_clicked(b):
    # Clear search form
    MANAGE_search_textbox.value = ""
    clear_officer_fields()
    MANAGE_status_msg.value = ""
    MANAGE_update_button.disabled = True

  MANAGE_clear_button.on_click(on_MANAGE_clear_button_clicked)

  def on_MANAGE_return_button_clicked(b):
    global refresh_view_officers

    # Reset the flag before navigating
    if refresh_view_officers:
      refresh_view_officers = False
      # Navigate back to View All Officers tab and trigger refresh
      showPage("Admin - View All Officers")
      # Directly trigger the refresh
      on_VIEW_all_button_clicked(None)
    else:
      # Just navigate without refreshing
      showPage("Admin - View All Officers")

  MANAGE_return_button.on_click(on_MANAGE_return_button_clicked)

  def on_MANAGE_back_button_clicked(b):
    showPage("Main Menu")

  MANAGE_back_button.on_click(on_MANAGE_back_button_clicked)

  # Style the buttons in Manage Officer Status tab
  MANAGE_update_button.style.button_color = '#4CAF50'  # Green
  MANAGE_clear_button.style.button_color = '#f0f0f0'  # Light gray
  MANAGE_back_button.style.button_color = '#f0f0f0'  # Light gray
  MANAGE_return_button.style.button_color = '#2196F3'  # Blue

  # Modified display for Manage Officer Status tab with consistent spacing
  title_section = widgets.VBox([
    MANAGE_title,
    MANAGE_description,
    SEPARATOR
  ])

  search_section = widgets.HBox([
    MANAGE_search_label,
    MANAGE_search_textbox,
    MANAGE_search_button
  ], layout=widgets.Layout(margin='10px 0', align_items='center'))

  MANAGE_labels = widgets.VBox([
    MANAGE_id_label,
    MANAGE_username_label,
    MANAGE_fname_label,
    MANAGE_lname_label,
    MANAGE_phone_label,
    MANAGE_email_label,
    MANAGE_role_label,
    MANAGE_status_label
  ], layout=widgets.Layout(justify_content='space-between', align_items='flex-end'))

  MANAGE_inputs = widgets.VBox([
    MANAGE_id_textbox,
    MANAGE_username_textbox,
    MANAGE_fname_textbox,
    MANAGE_lname_textbox,
    MANAGE_phone_textbox,
    MANAGE_email_textbox,
    MANAGE_role_textbox,
    MANAGE_status_dropdown
  ], layout=widgets.Layout(justify_content='space-between'))

  officer_details = widgets.HBox([MANAGE_labels, MANAGE_inputs],
                               layout=widgets.Layout(margin='15px 0', align_items='center'))

  # Create button row with spacing
  button_section = widgets.HBox([
    MANAGE_update_button,
    widgets.HTML("&nbsp;&nbsp;"),  # Add space between buttons
    MANAGE_clear_button,
    widgets.HTML("&nbsp;&nbsp;"),  # Add space between buttons
    MANAGE_return_button,
    widgets.HTML("&nbsp;&nbsp;"),  # Add space between buttons
    MANAGE_back_button
  ], layout=widgets.Layout(margin='15px 0', justify_content='flex-start'))

  display(widgets.VBox([
    title_section,
    search_section,
    MANAGE_spacer,
    officer_details,
    SECTION_SPACING,
    MANAGE_status_msg,
    SECTION_SPACING,
    button_section
  ], layout=widgets.Layout(padding='20px')))
##############################################################
########### Admin - Manage Officer Status Tab - End ##########
##############################################################

# Create a custom function to show the Manage Officer Status page with cleared form
def showManageOfficerStatusPage():
  # Clear all form fields first
  MANAGE_search_textbox.value = ""
  clear_officer_fields()
  MANAGE_status_msg.value = ""
  MANAGE_update_button.disabled = True

  # Then show the page
  showPage("Admin - Manage Officer Status")

##############################################################
########### Admin - Edit Products Tab - Begin ################
##############################################################
with tb_all.output_to('Admin - Edit Products'):
    # Header and description
    edit_title = create_header("Edit Products")
    edit_description = create_subheader("Search for a product and update its details")

    # Create search section
    EDIT_search_label = create_styled_label("Search Product:")
    EDIT_search_textbox = Text(value='', placeholder='Enter Product ID or Keyword', disabled=False)
    EDIT_search_button = widgets.Button(description="Search", layout=SEARCH_BUTTON_LAYOUT)
    style_button(EDIT_search_button, PRIMARY_COLOR)

    # Create display section
    EDIT_results_label = create_styled_label("Search Results:")
    EDIT_resultsremind_label = create_styled_label("Please select a product from the dropdown:")
    EDIT_results_dropdown = widgets.Dropdown(options=[], description='', disabled=False)

    # Create fields for editing product info
    EDIT_fields = {
        "id": Text(value='', placeholder='Product ID', disabled=True),
        "name": Text(value='', placeholder='Product Name', disabled=False),
        "description": Text(value='', placeholder='Product Description', disabled=False),
        "price": Text(value='', placeholder='Price', disabled=False),
        "weight": Text(value='', placeholder='Weight', disabled=False),
        "colour": Text(value='', placeholder='Colour', disabled=False),
        "capacity": Text(value='', placeholder='Capacity', disabled=False),
        "manufacturer": Text(value='', placeholder='Manufacturer', disabled=False),
        "status": widgets.Dropdown(options=['Active', 'Inactive'], value='Active', description='', disabled=False),
        "inventory": IntText(value=0, description='', disabled=False),
    }

    # Status messages
    EDIT_status_msg = HTML("")
    EDIT_search_msg = HTML("")

    # Create buttons
    EDIT_update_button = widgets.Button(description="Update Product", disabled=True, layout=ACTION_BUTTON_LAYOUT)
    style_button(EDIT_update_button, PRIMARY_COLOR)

    EDIT_clear_button = widgets.Button(description="Clear", layout=ACTION_BUTTON_LAYOUT)
    style_button(EDIT_clear_button, NEUTRAL_COLOR)

    EDIT_back_button = widgets.Button(description="Back to Menu", layout=ACTION_BUTTON_LAYOUT)
    style_button(EDIT_back_button, NEUTRAL_COLOR)

    # Clear function for resetting fields
    def clear_all_fields():
        EDIT_search_textbox.value = ""
        for field in EDIT_fields.values():
            if isinstance(field, Text):
                field.value = ""
            elif isinstance(field, IntText):
                field.value = 0
            elif isinstance(field, widgets.Dropdown):
                field.value = field.options[0]  # Set to the first option

        EDIT_status_msg.value = ""
        EDIT_results_dropdown.options = []
        EDIT_results_dropdown.value = None  # Reset dropdown value
        EDIT_update_button.disabled = True

    # Define button handlers
    def on_EDIT_search_button_clicked(b):
        try:
            search_value = EDIT_search_textbox.value.strip()
            query = """
                SELECT * FROM PRODUCT
                WHERE UPPER(P_NAME) LIKE UPPER(:keyword) OR
                      UPPER(P_DESCRIPTION) LIKE UPPER(:keyword) OR
                      P_ID = :id
            """
            # Prepare keyword for LIKE search
            keyword = '%' + search_value.upper() + '%'
            id_value = int(search_value) if search_value.isdigit() else None

            c.execute(query, {'keyword': keyword, 'id': id_value})
            rows = c.fetchall()

            if rows:
                EDIT_results_dropdown.options = [f"ID: {row[0]}, Name: {row[1]}" for row in rows]
                EDIT_search_msg.value = create_styled_message(f"{len(rows)} product(s) found.", "success").value
                EDIT_update_button.disabled = False
            else:
                clear_all_fields()
                EDIT_search_msg.value = create_styled_message("No products found.", "error").value

        except Exception as e:
            EDIT_search_msg.value = create_styled_message(f"Error searching for products: {str(e)}", "error").value

    EDIT_search_button.on_click(on_EDIT_search_button_clicked)

    # Populate fields when a product is selected
    def on_product_selected(change):
        if EDIT_results_dropdown.value:
            selected_id = EDIT_results_dropdown.value.split(',')[0].split(':')[1].strip()
            c.execute("SELECT * FROM PRODUCT WHERE P_ID = :id", {'id': selected_id})
            row = c.fetchone()
            if row:
                populate_product_fields(row)

    EDIT_results_dropdown.observe(on_product_selected, names='value')

    def populate_product_fields(row):
        EDIT_fields["id"].value = str(row[0])
        EDIT_fields["name"].value = row[1]
        EDIT_fields["description"].value = row[2]
        EDIT_fields["price"].value = row[3]
        EDIT_fields["weight"].value = row[4]
        EDIT_fields["colour"].value = row[5]
        EDIT_fields["capacity"].value = row[6]
        EDIT_fields["manufacturer"].value = row[7]
        EDIT_fields["status"].value = row[8]
        EDIT_fields["inventory"].value = row[9]

    def on_EDIT_update_button_clicked(b):
        try:
            EDIT_fields["status"].value = "Inactive" if EDIT_fields["inventory"].value == 0 else "Active"

            c.execute("""
                UPDATE PRODUCT
                SET P_NAME = :name,
                    P_DESCRIPTION = :description,
                    P_PRICE = :price,
                    P_WEIGHT = :weight,
                    P_COLOUR = :colour,
                    P_CAPACITY = :capacity,
                    P_MANUFACTURER = :manufacturer,
                    P_STATUS = :status,
                    P_INVENTORY = :inventory
                WHERE P_ID = :id
            """, {
                'name': EDIT_fields["name"].value,
                'description': EDIT_fields["description"].value,
                'price': EDIT_fields["price"].value,
                'weight': EDIT_fields["weight"].value,
                'colour': EDIT_fields["colour"].value,
                'capacity': EDIT_fields["capacity"].value,
                'manufacturer': EDIT_fields["manufacturer"].value,
                'status': EDIT_fields["status"].value,
                'inventory': EDIT_fields["inventory"].value,
                'id': EDIT_fields["id"].value
            })

            c.execute("commit")
            EDIT_status_msg.value = create_styled_message("Product updated successfully!", "success").value
        except Exception as e:
            EDIT_status_msg.value = create_styled_message(f"Error updating product: {str(e)}", "error").value

    EDIT_update_button.on_click(on_EDIT_update_button_clicked)

    def on_EDIT_clear_button_clicked(b):
        clear_all_fields()

    EDIT_clear_button.on_click(on_EDIT_clear_button_clicked)

    def on_EDIT_back_button_clicked(b):
        clear_all_fields()  # Clear fields and reset messages
        EDIT_search_msg.value = ""  # Clear search message
        showPage("Main Menu")

    EDIT_back_button.on_click(on_EDIT_back_button_clicked)

    # Display the objects
    edit_intro = widgets.VBox([edit_title, edit_description])
    EDIT_search_box = widgets.HBox([EDIT_search_label, EDIT_search_textbox, EDIT_search_button])
    EDIT_resultsmsg_box = widgets.HBox([EDIT_results_label, EDIT_search_msg])
    EDIT_results_box = widgets.VBox([EDIT_resultsremind_label, EDIT_results_dropdown])

    EDIT_labels = widgets.VBox([create_styled_label(label) for label in EDIT_fields.keys()])
    EDIT_inputs = widgets.VBox(list(EDIT_fields.values()))

    EDIT_form = widgets.HBox([EDIT_labels, EDIT_inputs])
    EDIT_buttons = widgets.HBox([EDIT_update_button, EDIT_clear_button, EDIT_back_button])

    display(widgets.VBox([edit_intro, SECTION_SPACING, EDIT_search_box, EDIT_resultsmsg_box, EDIT_results_box,
                          SECTION_SPACING, EDIT_form, EDIT_status_msg, EDIT_buttons]))
################ Admin - Edit Products Tab - End ################

##############################################################
########### Admin - View All Orders Tab - Begin #############
##############################################################
with tb_all.output_to('Admin - View All Orders'):

    # Create header with title using standard style
    VIEW_title = create_header("View All Orders")

    # Create retrieve all button with standard layout
    VIEW_all_button = widgets.Button(
        description="Load All Orders",
        layout=ACTION_BUTTON_LAYOUT
    )
    VIEW_back_button = widgets.Button(
        description="Back to Menu",
        layout=ACTION_BUTTON_LAYOUT
    )

    # Apply styles to buttons using the helper function
    style_button(VIEW_all_button, PRIMARY_COLOR)
    style_button(VIEW_back_button, NEUTRAL_COLOR)

    # Create output for displaying all orders
    VIEW_all_output = widgets.Output()

    # Create message for displaying number of retrieved orders
    VIEW_all_count_msg = Label("")
    VIEW_all_count_msg.style = STATUS_MSG_STYLE

    # Lists for storing textboxes when displaying all orders
    VIEW_all_id_textboxes = []
    VIEW_all_status_textboxes = []
    VIEW_all_date_textboxes = []
    VIEW_all_HBoxes = []

    # Status message for errors
    VIEW_status_msg = Label("")
    VIEW_status_msg.style = ERROR_STYLE

    def on_VIEW_all_button_clicked(b):
        try:
            # Clear previous data
            VIEW_all_id_textboxes.clear()
            VIEW_all_status_textboxes.clear()
            VIEW_all_date_textboxes.clear()
            VIEW_all_HBoxes.clear()

            VIEW_all_output.clear_output()
            VIEW_all_count_msg.value = ""
            VIEW_status_msg.value = ""

            # Retrieve all orders from database
            c.execute("""
                SELECT O_ID, O_STATUS, O_DATE FROM ORDERS
                ORDER BY O_ID ASC
            """)

            rows = c.fetchall()
            if not rows:
                VIEW_all_count_msg.value = "No orders found in the database."
                return

            # Define consistent widths for better alignment
            id_width = '60px'
            status_width = '100px'
            date_width = '100px'
            button_width = '80px'

            # Apply consistent styles to all cells
            header_style = {'font_weight': 'bold', 'background': '#f2f2f2', 'padding': '5px'}
            cell_style = {'padding': '3px'}

            # Create a header row
            header = widgets.HBox([
                Label("Order ID", layout=widgets.Layout(width=id_width), style=header_style),
                Label("Status", layout=widgets.Layout(width=status_width), style=header_style),
                Label("Date", layout=widgets.Layout(width=date_width), style=header_style),
                Label("Action", layout=widgets.Layout(width=button_width), style=header_style)
            ])

            # Display results with consistent widths
            for row in rows:
                # Create textboxes with consistent styling
                id_text = Text(value=str(row[0]), layout=widgets.Layout(width=id_width, padding='2px'), disabled=True)
                VIEW_all_id_textboxes.append(id_text)

                status_text = Text(value=str(row[1]), layout=widgets.Layout(width=status_width, padding='2px'), disabled=True)
                VIEW_all_status_textboxes.append(status_text)

                date_text = Text(value=str(row[2]), layout=widgets.Layout(width=date_width, padding='2px'), disabled=True)
                VIEW_all_date_textboxes.append(date_text)

                # Create action button for viewing order details (optional)
                view_btn = widgets.Button(
                    description="View",
                    button_style='info',
                    layout=widgets.Layout(width=button_width, padding='2px')
                )
                view_btn.order_id = str(row[0])  # Store order ID as attribute
                view_btn.on_click(on_view_order_button_clicked)

                # Create HBoxes for each order row
                row_box = widgets.HBox([
                    id_text,
                    status_text,
                    date_text,
                    view_btn
                ])
                VIEW_all_HBoxes.append(row_box)

            # Add a container for the table with a border
            table_container = widgets.VBox([header] + VIEW_all_HBoxes)
            table_container.layout.border = '1px solid #ddd'

            # Display all orders
            with VIEW_all_output:
                display(table_container)

            # Update count message
            VIEW_all_count_msg.value = f"Retrieved {len(VIEW_all_id_textboxes)} orders."

        except Exception as e:
            VIEW_status_msg.value = f"Error retrieving orders: {str(e)}"

    def on_view_order_button_clicked(b):
        order_id = b.order_id
        # Navigate to Manage Orders tab (or a details view)
        showPage("Admin - Manage Orders")
        # Pre-populate the search field with the order ID (if applicable)
        ORDER_search_textbox.value = order_id
        # Trigger the search button click to load order details (if applicable)
        on_ORDER_search_button_clicked(None)

    VIEW_all_button.on_click(on_VIEW_all_button_clicked)

    def on_VIEW_back_button_clicked(b):
        showPage("Main Menu")

    VIEW_back_button.on_click(on_VIEW_back_button_clicked)

    # Modified display for View All Orders tab with consistent spacing
    display(widgets.VBox([
        VIEW_title,
        widgets.HBox([VIEW_all_button, VIEW_back_button], layout=widgets.Layout(margin='10px 0')),
        SECTION_SPACING,
        VIEW_all_output,
        VIEW_all_count_msg,
        VIEW_status_msg
    ], layout=widgets.Layout(padding='20px')))
########### Admin - View All Orders Tab - End ################

##############################################################
########### Admin - Manage Orders Tab - Begin ################
##############################################################
with tb_all.output_to('Admin - Manage Orders'):
    # Header and description
    order_title = create_header("Manage Orders")
    order_description = create_subheader("Search for an order and update its status")

    # Create search section
    ORDER_search_label = create_styled_label("Search Order ID:")
    ORDER_search_textbox = Text(value='', placeholder='Enter Order ID', disabled=False)
    ORDER_search_button = widgets.Button(description="Search Order", layout=SEARCH_BUTTON_LAYOUT)
    style_button(ORDER_search_button, PRIMARY_COLOR)

    # Create display section
    ORDER_results_label = create_styled_label("Search Result:")
    ORDER_status_msg = HTML("")  # Use HTML to properly format the message

    # Fields to display order information
    ORDER_id_label = create_styled_label("Order ID:")
    ORDER_id_textbox = Text(value='', placeholder='Order ID', disabled=True)

    ORDER_status_label = create_styled_label("Current Status:")
    ORDER_status_dropdown = widgets.Dropdown(
        options=['Pending', 'To be delivered', 'Completed', 'Cancelled'],
        value='Pending',
        description='',
        disabled=False
    )

    # Status message for updates
    ORDER_update_msg = HTML("")

    # Create buttons
    ORDER_update_button = widgets.Button(description="Update Status", disabled=True, layout=ACTION_BUTTON_LAYOUT)
    style_button(ORDER_update_button, PRIMARY_COLOR)

    ORDER_clear_button = widgets.Button(description="Clear", layout=ACTION_BUTTON_LAYOUT)
    style_button(ORDER_clear_button, NEUTRAL_COLOR)

    ORDER_back_button = widgets.Button(description="Back to Menu", layout=ACTION_BUTTON_LAYOUT)
    style_button(ORDER_back_button, NEUTRAL_COLOR)

    # Clear function for resetting fields and messages
    def clear_all_order_fields():
        ORDER_search_textbox.value = ""
        ORDER_id_textbox.value = ""
        ORDER_status_dropdown.value = 'Pending'
        ORDER_update_button.disabled = True
        ORDER_status_msg.value = ""
        ORDER_update_msg.value = ""

    # Define button handlers
    def on_ORDER_search_button_clicked(b):
        try:
            order_id = ORDER_search_textbox.value.strip()

            # Check if order_id is numeric
            if not order_id.isdigit():
                ORDER_status_msg.value = create_styled_message("Please enter a numeric order ID!", "error").value
                clear_all_order_fields()  # Clear fields if invalid
                return

            c.execute("SELECT O_ID, O_STATUS FROM ORDERS WHERE O_ID = :id", {'id': order_id})
            row = c.fetchone()
            if row:
                ORDER_id_textbox.value = str(row[0])
                ORDER_status_dropdown.value = row[1]
                ORDER_update_button.disabled = (row[1] == 'Cancelled')  # Disable if status is Cancelled

                # Reminder message for cancelled orders
                if row[1] == 'Cancelled':
                    ORDER_status_msg.value = create_styled_message(
                        f"Order ID {order_id} found. Note: You cannot change the status of a Cancelled order.", "warning").value
                else:
                    ORDER_status_msg.value = create_styled_message(f"Order ID {order_id} found.", "success").value
            else:
                clear_all_order_fields()  # Clear fields if order not found
                ORDER_status_msg.value = create_styled_message("Order not found.", "error").value
        except Exception as e:
            ORDER_status_msg.value = create_styled_message(f"Error searching for order: {str(e)}", "error").value

    ORDER_search_button.on_click(on_ORDER_search_button_clicked)

    def on_ORDER_update_button_clicked(b):
        try:
            # Prevent updates if the status is Cancelled
            if ORDER_status_dropdown.value == 'Cancelled':
                ORDER_update_msg.value = create_styled_message("Cannot change the status of a Cancelled order.", "error").value
                return

            order_id = ORDER_id_textbox.value
            new_status = ORDER_status_dropdown.value
            c.execute("""
                UPDATE ORDERS
                SET O_STATUS = :status
                WHERE O_ID = :id
            """, {'status': new_status, 'id': order_id})
            c.execute("commit")
            ORDER_update_msg.value = create_styled_message("Order status updated successfully!", "success").value
        except Exception as e:
            ORDER_update_msg.value = create_styled_message(f"Error updating order status: {str(e)}", "error").value

    ORDER_update_button.on_click(on_ORDER_update_button_clicked)

    def on_ORDER_clear_button_clicked(b):
        clear_all_order_fields()  # Clear fields when clear button is clicked

    ORDER_clear_button.on_click(on_ORDER_clear_button_clicked)

    def on_ORDER_back_button_clicked(b):
        clear_all_order_fields()  # Clear fields and reset messages
        showPage("Main Menu")

    ORDER_back_button.on_click(on_ORDER_back_button_clicked)

    # Display the objects
    ORDER_intro = widgets.VBox([order_title, order_description])
    ORDER_search_box = widgets.HBox([ORDER_search_label, ORDER_search_textbox, ORDER_search_button])
    ORDER_resultmsg_box = widgets.HBox([ORDER_results_label, ORDER_status_msg])
    ORDER_details_box = widgets.VBox([ORDER_id_label, ORDER_id_textbox,
                                       ORDER_status_label, ORDER_status_dropdown, ORDER_update_msg])
    ORDER_buttons = widgets.HBox([ORDER_update_button, ORDER_clear_button, ORDER_back_button])

    display(widgets.VBox([ORDER_intro, SECTION_SPACING, ORDER_search_box, ORDER_resultmsg_box, ORDER_details_box, ORDER_buttons]))
################ Admin - Manage Orders Tab - End ################

##############################################################
########### Officer - Change Password Tab - Begin ############
##############################################################
with tb_all.output_to('Officer - Change Password'):
  # Header and description
  password_title = create_header("Change Password")
  password_description = create_subheader("Officier can change their own passwords")

  # Create labels and textboxes for the password change interface
  PWD_officer_id_label = create_styled_label("Officer ID: ")
  PWD_officer_id_textbox = Text(value='', placeholder='Enter Officer ID', layout=INPUT_LAYOUT, disabled=True)

  PWD_spacer = SEPARATOR

  PWD_current_pass_label = create_styled_label("Current Password: ")
  PWD_new_pass_label = create_styled_label("New Password: ")
  PWD_confirm_pass_label = create_styled_label("Confirmed Password: ")

  PWD_current_pass_textbox = Password(value='', placeholder='Enter Current Password', disabled=False)
  plain_text_currentpwd = Text(value='', placeholder='Enter Current Password', disabled=False)
  plain_text_currentpwd.layout.display = 'none'

  PWD_new_pass_textbox = Password(value='', placeholder='Enter New Password', disabled=False)
  plain_text_newpwd = Text(value='', placeholder='Enter New Password', disabled=False)
  plain_text_newpwd.layout.display = 'none'

  PWD_confirm_pass_textbox = Password(value='', placeholder='Enter New Password', disabled=False)
  plain_text_confirmpwd = Text(value='', placeholder='Enter New Password', disabled=False)
  plain_text_confirmpwd.layout.display = 'none'

  password_fields.extend([
    (PWD_current_pass_textbox, plain_text_currentpwd),
    (PWD_new_pass_textbox, plain_text_newpwd),
    (PWD_confirm_pass_textbox, plain_text_confirmpwd)])

  # Status message display
  PWD_status_label = HTML("")

  # Create buttons
  PWD_change_button = widgets.Button(description="Change Password", layout=ACTION_BUTTON_LAYOUT)
  PWD_clear_button = widgets.Button(description="Clear", layout=ACTION_BUTTON_LAYOUT)
  PWD_back_button = widgets.Button(description="Back to Menu", layout=ACTION_BUTTON_LAYOUT)

  # Apply consistent button styling
  style_button(PWD_change_button, PRIMARY_COLOR)
  style_button(PWD_clear_button, NEUTRAL_COLOR)
  style_button(PWD_back_button, NEUTRAL_COLOR)

  # Create checkbox
  cb2 = Checkbox(description='Show Password', value=False, disabled=False, indent=False)
  cb3 = Checkbox(description='Show Password', value=False, disabled=False, indent=False)
  cb4 = Checkbox(description='Show Password', value=False, disabled=False, indent=False)
  checkbox.extend([cb2, cb3, cb4])

  cb2.observe(on_show_password_checkbox_change, names='value')
  cb3.observe(on_show_password_checkbox_change, names='value')
  cb4.observe(on_show_password_checkbox_change, names='value')

  # Define the event handlers for the buttons when clicked
  def on_PWD_change_button_clicked(b):
    password_value = plain_text_currentpwd.value if cb2.value else PWD_current_pass_textbox.value
    password_value1 = plain_text_newpwd.value if cb3.value else PWD_new_pass_textbox.value
    password_value2 = plain_text_confirmpwd.value if cb4.value else PWD_confirm_pass_textbox.value

    # Use the current logged-in user's ID
    officer_id = current_user['id']

    # First, verify the current password
    c.execute("""
        SELECT * FROM STAFF
        WHERE S_ID = :staffid AND S_PW = :password
      """, {'staffid': officer_id, 'password': password_value})

    result = c.fetchone()

    if not result:
      PWD_status_label.value = create_styled_message("Current password is incorrect. Please try again.", "error").value
      return

    # Check if new password and confirm password match
    if password_value1 != password_value2:
      PWD_status_label.value = create_styled_message("New password and confirm password do not match. Please try again.", "error").value
      return

    #Check if the new password and current password are different
    if password_value == password_value2:
      PWD_status_label.value = create_styled_message("New password must be different from your current password. Please try again.", "error").value
      return

    # Update the password in the database
    try:
      c.execute("""
        UPDATE STAFF
        SET S_PW = :password
        WHERE S_ID =:staffid
      """, {'staffid': officer_id, 'password': password_value1})

      c.execute("commit")
      PWD_status_label.value = create_styled_message("Password updated successfully!", "success").value

    except Exception as e:
      PWD_status_label.value = create_styled_message("Error updating password: " + str(e), "error").value

  PWD_change_button.on_click(on_PWD_change_button_clicked)

  def on_PWD_clear_button_clicked(b):
    clear_fields(PWD_current_pass_textbox, PWD_new_pass_textbox, PWD_confirm_pass_textbox, plain_text_currentpwd,
                 plain_text_newpwd, plain_text_confirmpwd, PWD_status_label, cb2, cb3, cb4)
  PWD_clear_button.on_click(on_PWD_clear_button_clicked)

  def on_PWD_back_button_clicked(b):
    clear_fields(PWD_current_pass_textbox, PWD_new_pass_textbox, PWD_confirm_pass_textbox, plain_text_currentpwd,
                 plain_text_newpwd, plain_text_confirmpwd, PWD_status_label, cb2, cb3, cb4)
    showPage("Main Menu")

  PWD_back_button.on_click(on_PWD_back_button_clicked)

  # This function will be called when this tab is shown
  def setup_password_change_tab():
    PWD_officer_id_textbox.value = str(current_user['id'])

  # Display the objects
  password_current = widgets.HBox([PWD_current_pass_textbox, plain_text_currentpwd, cb2])
  password_new = widgets.HBox([PWD_new_pass_textbox, plain_text_newpwd, cb3])
  password_confirm = widgets.HBox([PWD_confirm_pass_textbox, plain_text_confirmpwd, cb4])

  password_intro = widgets.VBox([password_title, password_description])
  password_label = widgets.VBox([PWD_officer_id_label, PWD_current_pass_label, PWD_new_pass_label, PWD_confirm_pass_label])
  password_textbox = widgets.VBox([PWD_officer_id_textbox, password_current, password_new, password_confirm])

  password_section = widgets.HBox([password_label, password_textbox])

  display(widgets.VBox([
        password_intro,
        PWD_spacer,
        password_section,
        PWD_status_label,
        widgets.HBox([PWD_back_button, PWD_change_button, PWD_clear_button])
    ]))
################ Officer - Change Password Tab - End ################

##############################################################
########### Officer - View Trade-in Items Tab - Begin ########
##############################################################
with tb_all.output_to('Officer - View Trade-in Items'):
    # Header and description
    tradein_title = create_header("View Trade-in Items")
    tradein_description = create_subheader('''View the details of trade-in items by member ID and order number.
                                              After inspecting the device, the officer will input a brief condition,
                                              provide a trade-in offer, and update the trade-in status as either Accepted or Rejected.''')

    # Create search section
    search_tradein_label = create_styled_label("Please enter both Member ID and Order Number to search for trade-in items:")
    search_member_label = create_styled_label("Member ID: ")
    search_order_label = create_styled_label("Order Number: ")

    search_member_textbox = Text(value='', placeholder='Enter Member ID', layout=INPUT_LAYOUT, disabled=False)
    search_order_textbox = Text(value='', placeholder='Enter Order Number', layout=INPUT_LAYOUT, disabled=False)
    search_tradein_button = widgets.Button(description="Search", layout=SEARCH_BUTTON_LAYOUT)

    # Create display section
    search_tradein_spacer = SEPARATOR
    search_tradein_results_label = create_styled_label("Search Results:")
    search_tradein_resultsremind_label = create_styled_label("Please select an item from the dropdown:")
    search_tradein_results_dropdown = widgets.Dropdown(options=[], description='', layout=DROPDOWN_LAYOUT, disabled=False)

    # Create fields for viewing trade-in device info
    tradein_id_label = create_styled_label("Trade-in Device ID: ")
    tradein_id_textbox = Text(value='', placeholder='Trade-in Device ID', layout=INPUT_LAYOUT, disabled=True)

    tradein_serial_no_label = create_styled_label("Serial Number: ")
    tradein_serial_no_textbox = Text(value='', placeholder='Trade-in Device Serial Number', layout=INPUT_LAYOUT, disabled=True)

    tradein_condition_label = create_styled_label("Condition: ")
    tradein_condition_textbox = Text(value='', placeholder='Trade-in Device Condition', layout=INPUT_LAYOUT, disabled=False)

    tradein_price_label = create_styled_label("Trade-in Price: ")
    tradein_price_textbox = Text(value='', placeholder='Trade-in Price', layout=INPUT_LAYOUT, disabled=False)

    tradein_status_label = create_styled_label("Status: ")
    tradein_status_dropdown = widgets.Dropdown(
        options=['Requested', 'Accepted', 'Rejected'],
        value='Requested',
        description='',
        layout=DROPDOWN_LAYOUT,
        disabled=False)

    tradein_order_id_label = create_styled_label("Order ID: ")
    tradein_order_id_textbox = Text(value='', placeholder='Order ID', layout=INPUT_LAYOUT, disabled=True)

    tradein_staff_id_label = create_styled_label("Staff ID: ")
    tradein_staff_id_textbox = Text(value='', placeholder='Staff ID', layout=INPUT_LAYOUT, disabled=True)

    # Status message
    tradein_search_msg = HTML("")
    tradein_status_msg = HTML("")

    # Create buttons
    tradein_update_button = widgets.Button(description="Update", disabled = True,
                                              layout=ACTION_BUTTON_LAYOUT)
    tradein_clear_button = widgets.Button(description="Clear", layout=ACTION_BUTTON_LAYOUT)
    tradein_back_button = widgets.Button(description="Back to Menu", layout=ACTION_BUTTON_LAYOUT)
    tradein_process_button = widgets.Button(description="Process Order", disabled = True,
                                              layout=ACTION_BUTTON_LAYOUT)

    # Apply consistent button styling
    style_button(search_tradein_button, PRIMARY_COLOR)
    style_button(tradein_update_button, PRIMARY_COLOR)
    style_button(tradein_process_button, PRIMARY_COLOR)
    style_button(tradein_clear_button, NEUTRAL_COLOR)
    style_button(tradein_back_button, NEUTRAL_COLOR)

    # Define button handlers
    def on_search_tradein_button_clicked(b):
      try:
        # Reset
        search_tradein_results_dropdown.options = []
        clear_fields(tradein_status_msg)

        search_value = search_member_textbox.value.strip()
        search_value1 = search_order_textbox.value.strip()

        # Check for empty inputs
        if not search_value or not search_value1:
          tradein_search_msg.value = create_styled_message("Please input both Member ID and Order ID.", "error").value
          return

        # Check if both inputs are digits
        if search_value.isdigit() and search_value1.isdigit():
          c.execute("""
                SELECT T.T_ID, T.SERIAL_NUMBER
                FROM TRADE_IN_DEVICE T
                INNER JOIN ORDERS O ON T.ORDERS_O_ID = O.O_ID
                WHERE O.MEMBER_M_ID = :member_id AND O.O_ID = :order_id
            """, {'member_id': int(search_value), 'order_id': int(search_value1)})

          rows = c.fetchall()  # Fetch rows after executing the query

          if rows:
            # Populate the dropdown with results
            search_tradein_results_dropdown.options = [f"ID: {row[0]}, S/N: {row[1]}" for row in rows]
            tradein_search_msg.value = create_styled_message(f"{len(rows)} product(s) found.", "success").value
            tradein_update_button.disabled = False
          else:
            # Clear results if no products are found
            search_tradein_results_dropdown.options = []
            tradein_search_msg.value = create_styled_message("No trade-in devices found.", "error").value
            tradein_update_button.disabled = True
            tradein_process_button.disabled = True
            clear_fields(tradein_status_msg)

            # Reset fields
            tradein_status_dropdown.value = "Requested"
            clear_fields(tradein_id_textbox, tradein_serial_no_textbox, tradein_condition_textbox,
                          tradein_price_textbox, tradein_order_id_textbox, tradein_staff_id_textbox)
        else:
          tradein_search_msg.value = create_styled_message("Please enter valid numeric values for both Member ID and Order ID.", "error").value

      except Exception as e:
        tradein_search_msg.value = create_styled_message(f"Error searching for products: {str(e)}", "error").value

    search_tradein_button.on_click(on_search_tradein_button_clicked)

    # Populate fields when a trade-in device is selected
    def on_tradein_device_selected(change):
        clear_fields(tradein_status_msg)

        if search_tradein_results_dropdown.value:
            selected_id = search_tradein_results_dropdown.value.split(',')[0].split(':')[1].strip()
            c.execute("SELECT * FROM TRADE_IN_DEVICE WHERE T_ID = :id", {'id': selected_id})
            row = c.fetchone()
            if row:
                tradein_id_textbox.value = str(row[0])
                tradein_serial_no_textbox.value = row[1]
                tradein_condition_textbox.value = row[2]
                tradein_price_textbox.value = str(row[3])
                tradein_status_dropdown.value = row[4] if row[4] in tradein_status_dropdown.options else tradein_status_dropdown.options[0]
                tradein_order_id_textbox.value = str(row[5])
                tradein_staff_id_textbox.value = str(row[6])
                tradein_process_button.disabled = tradein_status_dropdown.value == 'Requested'
            else:
              # Clear fields if no selection
              clear_fields(tradein_id_textbox, tradein_serial_no_textbox, tradein_condition_textbox,
                     tradein_price_textbox, tradein_order_id_textbox, tradein_staff_id_textbox)

    search_tradein_results_dropdown.observe(on_tradein_device_selected, names='value')

    def on_tradein_update_button_clicked(b):
      try:
        c.execute("""
                UPDATE TRADE_IN_DEVICE
                SET SERIAL_NUMBER = :serial_no,
                    CONDITION = :condition,
                    TRADE_IN_PRICE = :price,
                    T_STATUS = :status,
                    ORDERS_O_ID = :orderID,
                    STAFF_S_ID = :staffID
                WHERE T_ID = :id
            """, {
                'serial_no': tradein_serial_no_textbox.value,
                'condition': tradein_condition_textbox.value,
                'price': float(tradein_price_textbox.value) if tradein_status_dropdown.value != 'Rejected' else 0,
                'status': tradein_status_dropdown.value,
                'orderID': int(tradein_order_id_textbox.value),
                'staffID': int(tradein_staff_id_textbox.value),
                'id': tradein_id_textbox.value
            })
        # Commit the changes
        c.execute("commit")
        tradein_status_msg.value = create_styled_message("Trade-in price and status updated successfully!", "success").value

        # Enable process button only if status is Accepted or Rejected
        tradein_process_button.disabled = tradein_status_dropdown.value != 'Accepted' and tradein_status_dropdown.value != 'Rejected'


      except Exception as e:
        tradein_status_msg.value = create_styled_message(f"Error updating status: {str(e)}", "error").value

    tradein_update_button.on_click(on_tradein_update_button_clicked)

    def on_tradein_process_button_clicked(b):
      process_order_id = tradein_order_id_textbox.value
      process_order(process_order_id)
    tradein_process_button.on_click(on_tradein_process_button_clicked)

    def process_order(order_id):
      # Set the order ID in the search text box
      process_search_textbox.value = order_id  # Assuming this is accessible globally
      showPage("Officer - Process Order")

    def on_tradein_clear_button_clicked(b):
      clear_fields(search_member_textbox, search_order_textbox, tradein_id_textbox, tradein_serial_no_textbox, tradein_condition_textbox,
                   tradein_price_textbox, tradein_order_id_textbox, tradein_staff_id_textbox, tradein_search_msg,
                   tradein_status_msg)

      tradein_status_dropdown.value = "Requested"
      search_tradein_results_dropdown.options = []
      tradein_update_button.disabled = True
      tradein_process_button.disabled = True

    tradein_clear_button.on_click(on_tradein_clear_button_clicked)

    def on_tradein_back_button_clicked(b):
      clear_fields(search_member_textbox, search_order_textbox, tradein_id_textbox, tradein_serial_no_textbox, tradein_condition_textbox,
                   tradein_price_textbox, tradein_order_id_textbox, tradein_staff_id_textbox, tradein_search_msg,
                   tradein_status_msg)

      tradein_status_dropdown.value = "Requested"
      search_tradein_results_dropdown.options = []
      tradein_update_button.disabled = True
      tradein_process_button.disabled = True
      showPage("Main Menu")

    tradein_back_button.on_click(on_tradein_back_button_clicked)

    # Display the objects
    tradein_intro = widgets.VBox([tradein_title, tradein_description])

    tradein_search_box = widgets.VBox([search_member_textbox, search_order_textbox])
    tradein_search_label = widgets.VBox([search_member_label, search_order_label])
    tradein_search_part = widgets.VBox([search_tradein_label, widgets.HBox([tradein_search_label,tradein_search_box]),search_tradein_button])

    tradein_resultsmsg_box = widgets.HBox([search_tradein_results_label, tradein_search_msg])
    tradein_results_box = widgets.VBox([tradein_resultsmsg_box, search_tradein_resultsremind_label, search_tradein_results_dropdown])
    tradein_labels = widgets.VBox([tradein_id_label, tradein_serial_no_label, tradein_condition_label,
                                tradein_price_label, tradein_status_label,
                                tradein_order_id_label, tradein_staff_id_label])
    tradein_inputs = widgets.VBox([tradein_id_textbox, tradein_serial_no_textbox, tradein_condition_textbox,
                                tradein_price_textbox, tradein_status_dropdown,
                                tradein_order_id_textbox, tradein_staff_id_textbox])

    tradein_form = widgets.HBox([tradein_labels, tradein_inputs])
    tradein_buttons = widgets.HBox([tradein_back_button, tradein_update_button, tradein_process_button, tradein_clear_button])

    display(widgets.VBox([tradein_intro, search_tradein_spacer, tradein_search_part,
                          search_tradein_spacer, tradein_results_box, tradein_form,
                          tradein_status_msg, tradein_buttons]))
################ Officer - View Trade-in Items - End ################

##############################################################
########### Officer - Process Order Tab - Begin ##############
##############################################################
with tb_all.output_to('Officer - Process Order'):

  # Header and description
  process_title = create_header("Process Order")
  process_description = create_subheader("Search for an order and update their status")

  # Create search section with better styling
  process_search_label = create_styled_label("Order ID: ")
  process_search_textbox = Text(value='', placeholder='Enter Order ID', layout=INPUT_LAYOUT, disabled=False)

  process_search_button = widgets.Button(
    description = "Search Order",
    layout=SEARCH_BUTTON_LAYOUT)
  style_button(process_search_button, PRIMARY_COLOR)

  # Status message for search and update
  process_search_msg = HTML("")
  process_status_msg = HTML("")

  # Create display section with separator
  process_spacer = SEPARATOR

  # Create fields for viewing officer info with consistent styling
  process_id_label = create_styled_label("Order ID: ")
  process_id_textbox = Text(value='', placeholder='Order ID', layout=INPUT_LAYOUT, disabled=True)

  process_amount_label = create_styled_label("Total Price (Before Trade-in Discount): ")
  process_amount_textbox = Text(value='', placeholder='Total prices of the order before trade-in discount', layout=INPUT_LAYOUT, disabled=True)

  process_final_amount_label = create_styled_label("Total Price (After Trade-in Discount): ")
  process_final_amount_textbox = Text(value='', placeholder='Total prices of the order included trade-in discount', layout=INPUT_LAYOUT, disabled=True)

  process_tradein_price_label = create_styled_label("Trade-in Discount: ")
  process_tradein_price_textbox = Text(value='', placeholder='Trade-in discount', layout=INPUT_LAYOUT, disabled=True)

  process_date_label = create_styled_label("Order Date: ")
  process_date_textbox = Text(value='', placeholder='Date of the order', layout=INPUT_LAYOUT, disabled=True)

  process_status_label = create_styled_label("Order Status: ")
  process_status_dropdown = widgets.Dropdown(
    options=['Pending','To be delivered', 'Cancelled'],
    value='Pending',
    description='',
    layout=DROPDOWN_LAYOUT,
    disabled=False)

  process_remark_label = create_styled_label("Remark: ")
  process_remark_textbox = Text(value='', placeholder='Additional remarks or comments about the order', layout=INPUT_LAYOUT, disabled=True)

  process_M_ID_label = create_styled_label("Member ID: ")
  process_M_ID_textbox = Text(value='', placeholder='Member ID', layout=INPUT_LAYOUT, disabled=True)

  # Create buttons with standardized layout and styling
  process_update_button = widgets.Button(description="Update Status", disabled=True, layout=ACTION_BUTTON_LAYOUT)
  process_clear_button = widgets.Button(description="Clear", layout=ACTION_BUTTON_LAYOUT)
  process_back_button = widgets.Button(description="Back to Menu", layout=ACTION_BUTTON_LAYOUT)

  # Apply consistent button styling
  style_button(process_update_button, PRIMARY_COLOR)
  style_button(process_clear_button, NEUTRAL_COLOR)
  style_button(process_back_button, NEUTRAL_COLOR)

  # Define button handlers
  def on_process_search_button_clicked(b):
    try:
      search_value = process_search_textbox.value.strip()
      clear_fields(process_status_msg)

      # Check for empty inputs
      if not search_value:
        process_search_msg.value = create_styled_message("Please input both Member ID and Order ID.", "error").value
        return

      # Check if both inputs are digits
      if search_value.isdigit():
        c.execute("""
              SELECT O.O_ID, O.O_DATE, O.O_TOTAL_AMOUNT,
              COALESCE(SUM(T.TRADE_IN_PRICE), 0) AS TOTAL_TRADE_IN_PRICE,
              O.O_STATUS, O.REMARKS, O.MEMBER_M_ID
              FROM ORDERS O
              INNER JOIN TRADE_IN_DEVICE T ON O.O_ID = T.ORDERS_O_ID
              WHERE O.O_ID = :id
              GROUP BY O.O_ID, O.O_DATE, O.O_TOTAL_AMOUNT, O.O_STATUS, O.REMARKS, O.MEMBER_M_ID
              """, {'id': process_search_textbox.value})

        row = c.fetchone()
        if row:
          process_id_textbox.value = str(row[0])
          process_date_textbox.value = str(row[1])
          process_amount_textbox.value = str(row[2])
          process_tradein_price_textbox.value = str(row[3])
          process_status_dropdown.value = row[4]
          process_remark_textbox.value = row[5]
          process_M_ID_textbox.value = str(row[6])
          process_final_amount_textbox.value = str(int(float(process_amount_textbox.value)) - int(float(process_tradein_price_textbox.value)))

          process_update_button.disabled = False
          process_search_msg.value = create_styled_message("Order found!", "success").value
        else:
          process_search_msg.value = create_styled_message("No order found with that ID. Please check the ID and try again.", "error").value
          clear_fields(process_id_textbox, process_amount_textbox, process_date_textbox,
                        process_remark_textbox, process_M_ID_textbox, process_status_msg, process_final_amount_textbox)
          process_status_dropdown.value = "Pending"
          process_update_button.disabled = True

      else:
        process_search_msg.value = create_styled_message("Please enter valid numeric values for Order ID.", "error").value

    except Exception as e:
      process_search_msg.value = create_styled_message(f"Error searching for order: {str(e)}", "error").value
      process_update_button.disabled = True

  process_search_button.on_click(on_process_search_button_clicked)

  def on_process_update_button_clicked(b):
    try:
      # Fetch the order lines for the current order
        c.execute("""
            SELECT OL.OL_QTY, OL.PRODUCT_P_ID
            FROM ORDER_LINE OL
            WHERE OL.ORDERS_O_ID = :id
           """, {'id': process_id_textbox.value})

        order_lines = c.fetchall()
        if process_status_dropdown.value == "Pending":
            process_status_msg.value = create_styled_message("No changes detected. Please try again.", "error").value

        elif process_status_dropdown.value == "Cancelled":
          # Update the inventory for each product in the order
          for qty, product_id in order_lines:
            c.execute("""
                UPDATE PRODUCT
                SET P_INVENTORY = P_INVENTORY + :qty
                WHERE P_ID = :product_id
              """, {'qty': qty, 'product_id': product_id})

            # Check the inventory after updating
            c.execute("""
                    SELECT P_INVENTORY
                    FROM PRODUCT
                    WHERE P_ID = :product_id
                """, {'product_id': product_id})

            current_inventory = c.fetchone()[0]

            # If inventory is greater than 0, change product status to 'Active'
            if current_inventory > 0:
              c.execute("""
                        UPDATE PRODUCT
                        SET P_STATUS = 'Active'
                        WHERE P_ID = :product_id
                    """, {'product_id': product_id})
          process_status_msg.value = create_styled_message(f"Order status updated successfully to {process_status_dropdown.value}!", "success").value

        elif process_status_dropdown.value == "To be delivered":
            # Update the total amount in the database
            new_total_amount = float(process_final_amount_textbox.value)
            c.execute("""
                UPDATE ORDERS
                SET O_TOTAL_AMOUNT = :total_amount
                WHERE O_ID = :id
            """, {'total_amount': new_total_amount, 'id': process_id_textbox.value})
            process_status_msg.value = create_styled_message(f"Order status updated successfully to {process_status_dropdown.value}!", "success").value

        # Update the order status
        c.execute("""
            UPDATE ORDERS
            SET O_STATUS = :status
            WHERE O_ID = :id
        """, {'status': process_status_dropdown.value, 'id': process_id_textbox.value})

        # Commit the changes
        c.execute("commit")

    except Exception as e:
      process_status_msg.value = create_styled_message(f"Error updating order status: {str(e)}", "error").value

  process_update_button.on_click(on_process_update_button_clicked)

  def on_process_clear_button_clicked(b):
    clear_fields(process_id_textbox, process_amount_textbox, process_date_textbox,
                  process_remark_textbox, process_M_ID_textbox, process_search_textbox,
                  process_status_msg, process_search_msg, process_tradein_price_textbox, process_final_amount_textbox)

    process_status_dropdown.value = "Pending"
    process_update_button.disabled = True

  process_clear_button.on_click(on_process_clear_button_clicked)

  def on_process_back_button_clicked(b):
      clear_fields(process_id_textbox, process_amount_textbox, process_date_textbox,
                  process_remark_textbox, process_M_ID_textbox, process_search_textbox,
                  process_status_msg, process_search_msg, process_tradein_price_textbox, process_final_amount_textbox)

      process_status_dropdown.value = "Pending"
      process_update_button.disabled = True
      showPage("Main Menu")

  process_back_button.on_click(on_process_back_button_clicked)

  # Display the objects
  process_title_box = widgets.VBox([process_title, process_description])
  process_search_box = widgets.HBox([process_search_label, process_search_textbox, process_search_button])

  process_labels = widgets.VBox([process_id_label, process_date_label, process_amount_label, process_tradein_price_label,
                                   process_final_amount_label, process_status_label, process_remark_label, process_M_ID_label])
  process_inputs = widgets.VBox([process_id_textbox, process_date_textbox, process_amount_textbox, process_tradein_price_textbox,
                                   process_final_amount_textbox, process_status_dropdown, process_remark_textbox, process_M_ID_textbox])

  process_form = widgets.HBox([process_labels, process_inputs])
  process_buttons = widgets.HBox([process_back_button, process_update_button, process_clear_button])

  display(widgets.VBox([process_title_box, process_spacer, process_search_box, process_search_msg,
                          process_spacer, process_form, process_status_msg,
                          process_buttons]))
################ Officer - Process Order Tab - End ###########


################## Manager Dashboard Tab ###########################

# Add this at the beginning of your code (with other global variables)
global download_button_exists
download_button_exists = False

# Modify the create_download_link function
def create_download_link(df, filename, button_text='Download Excel Report'):
    """Generate a download link for a pandas DataFrame as an Excel file with enhanced styling."""
    global download_button_exists

    # Clear any existing download button
    if download_button_exists:
        clear_output(wait=True)

    output = io.BytesIO()

    # Use openpyxl engine instead of xlsxwriter
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Write DataFrame to Excel
        df.to_excel(writer, sheet_name='Report', index=False, startrow=1, startcol=0)

        # Access the worksheet to apply styling
        worksheet = writer.sheets['Report']
        workbook = writer.book

        # Create a title for the report
        report_title = filename.replace('.xlsx', '').replace('_', ' ').title()
        worksheet.merge_cells('A1:F1')
        title_cell = worksheet['A1']
        title_cell.value = report_title

        # Create styles
        # Title style
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        title_font = Font(name='Calibri', size=16, bold=True, color='002060')
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Header style
        header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Apply header styles to the header row (row 2)
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=2, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Table borders
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        # Apply borders and styles to all data cells
        for row_num in range(2, len(df) + 3):  # +3 because of title row and header row
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.border = thin_border
                # Align text in cells
                if 'ID' in df.columns[col_num-1] or 'Count' in df.columns[col_num-1] or 'Quantity' in df.columns[col_num-1]:
                    cell.alignment = Alignment(horizontal='center')
                elif 'Price' in df.columns[col_num-1] or 'Amount' in df.columns[col_num-1] or 'Revenue' in df.columns[col_num-1] or 'Discount' in df.columns[col_num-1]:
                    cell.alignment = Alignment(horizontal='right')
                    # Format as currency if the cell contains price/amount data
                    if row_num > 2:  # Skip header row
                        # Check if value is a number that can be formatted
                        try:
                            if isinstance(cell.value, (int, float)):
                                cell.number_format = '$#,##0.00'
                        except:
                            pass
                else:
                    cell.alignment = Alignment(horizontal='left')

        # Alternate row colors for better readability
        for row_num in range(3, len(df) + 3):  # Start after header row
            if row_num % 2 == 0:
                for col_num in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.fill = PatternFill(start_color='E6F0F8', end_color='E6F0F8', fill_type='solid')

        # Auto-adjust column width - Fixed to avoid MergedCell error
        for col_num, column in enumerate(df.columns, 1):
            column_letter = get_column_letter(col_num)
            max_length = max(len(str(column)), 10)  # Set a minimum width based on column name

            # Check the data in each cell of the column
            for row_num in range(3, len(df) + 3):  # Skip header row
                try:
                    cell_value = worksheet.cell(row=row_num, column=col_num).value
                    if cell_value is not None:
                        max_length = max(max_length, len(str(cell_value)))
                except:
                    pass

            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Add summary information at the bottom if applicable
        if 'Total' in filename.lower() or 'Price' in filename.lower() or 'Revenue' in filename.lower():
            summary_row = len(df) + 4
            worksheet.merge_cells(f'A{summary_row}:D{summary_row}')
            summary_cell = worksheet[f'A{summary_row}']
            summary_cell.value = "Report Summary"
            summary_cell.font = Font(name='Calibri', size=12, bold=True)

            # Add total rows based on the type of report
            if 'Revenue' in filename:
                total_revenue = sum(df[df.columns[1]].astype(float)) if len(df) > 0 else 0
                worksheet[f'A{summary_row+1}'].value = "Total Revenue:"
                worksheet[f'B{summary_row+1}'].value = total_revenue
                worksheet[f'B{summary_row+1}'].number_format = '$#,##0.00'
                worksheet[f'B{summary_row+1}'].font = Font(bold=True)
            elif 'Quantity' in filename:
                total_qty = sum(df[df.columns[1]].astype(int)) if len(df) > 0 else 0
                worksheet[f'A{summary_row+1}'].value = "Total Quantity:"
                worksheet[f'B{summary_row+1}'].value = total_qty
                worksheet[f'B{summary_row+1}'].font = Font(bold=True)

    excel_data = output.getvalue()
    b64 = base64.b64encode(excel_data).decode()
    href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{filename}" class="excel-download-btn">{button_text}</a>'

    # Set the flag to indicate a download button exists
    download_button_exists = True

    return DisplayHTML(f"""
    <div style="margin: 10px 0;">
        {href}
        <style>
            .excel-download-btn {{
                display: inline-block;
                background-color: #2196F3;
                color: white;
                padding: 8px 15px;
                border-radius: 4px;
                text-decoration: none;
                font-weight: bold;
                margin: 10px 0;
                transition: background-color 0.3s;
            }}
            .excel-download-btn:hover {{
                background-color: #1976D2;
            }}
        </style>
    </div>
    """)

# Helper function to create a stable download area and display download link
def display_excel_download(download_area, df, filename, button_text="Download Excel Report"):
    """Helper function to display Excel download link in a stable, fixed download area."""
    with download_area:
        clear_output(wait=True)  # Use wait=True to reduce flickering
        download_link = create_download_link(
            df,
            filename,
            button_text
        )
        display(download_link)

with tb_all.output_to('Manager Dashboard'):
    # Title
    DASH_title = widgets.Label("Manager Dashboard")
    DASH_title.style = HEADER_STYLE

    # Add a dedicated download area for dashboard exports
    DASH_download_area = widgets.Output(layout={'height': '80px', 'min_height': '80px', 'margin': '10px 0', 'overflow': 'hidden'})

    # Card style for report sections
    CARD_STYLE = {
        'border': '1px solid #ddd',
        'border_radius': '5px',
        'padding': '20px',
        'margin': '15px 0',
        'background': '#ffffff',
        'min_width': '350px'
    }

    # Common table styles
    TABLE_HEADER_STYLE = {'font_weight': 'bold', 'font_size': '18px', 'text_align': 'left'}
    TABLE_DATA_STYLE = {'font_size': '16px'}
    TABLE_ROW_LAYOUT = {'border': '1px solid #ddd', 'padding': '10px'}
    TABLE_HEADER_LAYOUT = {'border': '1px solid #ddd', 'padding': '10px', 'background': '#e0e0e0'}

    # R1: Number of orders received today
    DASH_r1_label = widgets.Label("Orders Received Today:")
    DASH_r1_label.style = SUBHEADER_STYLE
    DASH_r1_value = widgets.Label("", style={'font_size': '18px', 'color': '#333', 'margin': '10px 0'})
    DASH_r1_button = widgets.Button(
        description="View All Orders",
        layout={'width': '200px', 'margin': '10px 0'},
        tooltip="View details of today's orders"
    )
    style_button(DASH_r1_button, PRIMARY_COLOR)

    # R2: Revenue of the current calendar month
    DASH_r2_label = widgets.Label("Revenue This Month:")
    DASH_r2_label.style = SUBHEADER_STYLE
    DASH_r2_value = widgets.Label("", style={'font_size': '18px', 'color': '#333', 'margin': '10px 0'})

    # R3: Top 5 popular items by number of sales
    DASH_r3_label = widgets.Label("Top 5 Popular Items by Sales:")
    DASH_r3_label.style = SUBHEADER_STYLE
    DASH_r3_output = widgets.Output()

    # R4: Sales revenue for the day
    DASH_r4_label = widgets.Label("Sales Revenue Today:")
    DASH_r4_label.style = SUBHEADER_STYLE
    DASH_r4_value = widgets.Label("", style={'font_size': '18px', 'color': '#333', 'margin': '10px 0'})

    # Buttons for reports with tooltips
    DASH_r5_button = widgets.Button(
        description="Sales Revenue by Date",
        layout={'width': '220px', 'margin': '5px'},
        tooltip="View total revenue per day"
    )
    DASH_r6_button = widgets.Button(
        description="Sales Quantity by Date",
        layout={'width': '220px', 'margin': '5px'},
        tooltip="View total units sold per day"
    )
    DASH_r7_button = widgets.Button(
        description="Trade-in Discounts by Date",
        layout={'width': '220px', 'margin': '5px'},
        tooltip="View trade-in discounts per day"
    )
    DASH_r8_button = widgets.Button(
        description="Old Devices Received",
        layout={'width': '220px', 'margin': '5px'},
        tooltip="View details of trade-in devices"
    )
    DASH_r9_button = widgets.Button(
        description="Out-of-Stock Products",
        layout={'width': '220px', 'margin': '5px'},
        tooltip="View products with zero inventory"
    )
    DASH_r10_button = widgets.Button(
        description="Top 10 Customers",
        layout={'width': '220px', 'margin': '5px'},
        tooltip="View customers with the most orders"
    )

    DASH_back_button = widgets.Button(
        description="Back to Menu",
        layout={'width': '220px', 'margin': '10px 0'},
        tooltip="Return to the main menu"
    )

    # Style all buttons
    for btn in [DASH_r5_button, DASH_r6_button, DASH_r7_button, DASH_r8_button, DASH_r9_button, DASH_r10_button]:
        style_button(btn, SECONDARY_COLOR)
    style_button(DASH_back_button, NEUTRAL_COLOR)

    # Function to refresh dashboard data
    def refresh_dashboard():
        try:
            # R1: Number of orders today
            c.execute("""
                SELECT COUNT(*)
                FROM ORDERS
                WHERE TO_CHAR(O_DATE, 'YYYY-MM-DD') = TO_CHAR(SYSDATE, 'YYYY-MM-DD')
            """)
            r1_count = c.fetchone()[0]
            DASH_r1_value.value = str(r1_count)

            # R2: Revenue this month
            c.execute("""
                SELECT SUM(O_TOTAL_AMOUNT)
                FROM ORDERS
                WHERE TO_CHAR(O_DATE, 'YYYY-MM') = TO_CHAR(SYSDATE, 'YYYY-MM')
            """)
            r2_total = c.fetchone()[0] or 0
            DASH_r2_value.value = f"${r2_total:,.2f}"

            # R3: Top 5 popular items
            DASH_r3_output.clear_output()
            c.execute("""
                SELECT P.P_NAME, SUM(OL.OL_QTY) as total_qty
                FROM ORDER_LINE OL
                INNER JOIN PRODUCT P ON OL.PRODUCT_P_ID = P.P_ID
                GROUP BY P.P_NAME
                ORDER BY total_qty DESC
                FETCH FIRST 5 ROWS ONLY
            """)
            r3_rows = c.fetchall()

            # Create a DataFrame for Excel export
            global r3_df
            if r3_rows:
                r3_df = pd.DataFrame(r3_rows, columns=["Product Name", "Units Sold"])
            else:
                r3_df = pd.DataFrame([], columns=["Product Name", "Units Sold"])

            with DASH_r3_output:
                clear_output(wait=True)
                if not r3_rows:
                    display(widgets.Label("No sales data available."))
                else:
                    # Create container for top 5 table
                    top5_container = widgets.VBox([])

                    # Add header
                    header = widgets.HBox([
                        widgets.Label("Product Name", style=TABLE_HEADER_STYLE, layout={'width': '400px'}),
                        widgets.Label("Units Sold", style={**TABLE_HEADER_STYLE, 'text_align': 'right'}, layout={'width': '150px'})
                    ], layout=TABLE_HEADER_LAYOUT)

                    rows_list = [header]

                    # Create rows for each product
                    for i, row in enumerate(r3_rows):
                        bg_color = '#f9f9f9' if i % 2 == 0 else '#ffffff'
                        product_row = widgets.HBox([
                            widgets.Label(row[0], style=TABLE_DATA_STYLE, layout={'width': '400px'}),
                            widgets.Label(str(row[1]), style={**TABLE_DATA_STYLE, 'text_align': 'right'}, layout={'width': '150px'})
                        ], layout={**TABLE_ROW_LAYOUT, 'background': bg_color})
                        rows_list.append(product_row)

                    # If fewer than 5 items, add empty rows to maintain consistent height
                    for i in range(len(r3_rows), 5):
                        bg_color = '#f9f9f9' if i % 2 == 0 else '#ffffff'
                        empty_row = widgets.HBox([
                            widgets.Label("", style=TABLE_DATA_STYLE, layout={'width': '400px'}),
                            widgets.Label("", style={**TABLE_DATA_STYLE, 'text_align': 'right'}, layout={'width': '150px'})
                        ], layout={**TABLE_ROW_LAYOUT, 'background': bg_color})
                        rows_list.append(empty_row)

                    # Add all rows to container
                    top5_container.children = rows_list
                    display(top5_container)

                    # Add Excel export button
                    r3_excel_button = widgets.Button(
                        description="Export to Excel",
                        layout={'width': '150px', 'margin': '10px 0'},
                        tooltip="Export Top 5 Popular Items to Excel"
                    )
                    style_button(r3_excel_button, SECONDARY_COLOR)

                    def on_r3_excel_button_clicked(b):
                        # Use the dedicated download area
                        display_excel_download(
                            DASH_download_area,
                            r3_df,
                            "top_5_popular_items.xlsx",
                            "Download Top 5 Items Report"
                        )

                    r3_excel_button.on_click(on_r3_excel_button_clicked)
                    display(r3_excel_button)

            # R4: Sales revenue today
            c.execute("""
                SELECT SUM(O_TOTAL_AMOUNT)
                FROM ORDERS
                WHERE TO_CHAR(O_DATE, 'YYYY-MM-DD') = TO_CHAR(SYSDATE, 'YYYY-MM-DD')
            """)
            r4_total = c.fetchone()[0] or 0
            DASH_r4_value.value = f"${r4_total:,.2f}"

        except Exception as e:
            DASH_r1_value.value = f"Error: {str(e)}"
            DASH_r2_value.value = f"Error: {str(e)}"
            DASH_r4_value.value = f"Error: {str(e)}"
            with DASH_r3_output:
                clear_output()
                display(widgets.Label(f"Error: {str(e)}"))

    # R1: View all orders (navigate to new tab)
    def on_r1_button_clicked(b):
        showPage("View All Orders")
        refresh_orders_page()

    DASH_r1_button.on_click(on_r1_button_clicked)

    # R5: Sales Revenue by Date (navigate to new tab)
    def on_r5_button_clicked(b):
        showPage("Sales Revenue by Date")
        refresh_r5_page()

    DASH_r5_button.on_click(on_r5_button_clicked)

    # R6: Sales Quantity by Date (navigate to new tab)
    def on_r6_button_clicked(b):
        showPage("Sales Quantity by Date")
        refresh_r6_page()

    DASH_r6_button.on_click(on_r6_button_clicked)

    # R7: Trade-in Discounts by Date (navigate to new tab)
    def on_r7_button_clicked(b):
        showPage("Trade-in Discounts by Date")
        refresh_r7_page()

    DASH_r7_button.on_click(on_r7_button_clicked)

    # R8: Old Devices Received (navigate to new tab)
    def on_r8_button_clicked(b):
        showPage("Old Devices Received")
        refresh_r8_page()

    DASH_r8_button.on_click(on_r8_button_clicked)

    # R9: Out-of-Stock Products (navigate to new tab)
    def on_r9_button_clicked(b):
        showPage("Out-of-Stock Products")
        refresh_r9_page()

    DASH_r9_button.on_click(on_r9_button_clicked)

    # R10: Top 10 Customers (navigate to new tab)
    def on_r10_button_clicked(b):
        showPage("Top 10 Customers")
        refresh_r10_page()

    DASH_r10_button.on_click(on_r10_button_clicked)

    # Back button
    def on_back_button_clicked(b):
        showPage("Main Menu")
    DASH_back_button.on_click(on_back_button_clicked)

    # Layout
    r1_card = widgets.VBox([
        DASH_r1_label,
        DASH_r1_value,
        DASH_r1_button
    ], layout=CARD_STYLE)

    r2_card = widgets.VBox([
        DASH_r2_label,
        DASH_r2_value
    ], layout=CARD_STYLE)

    r3_card = widgets.VBox([
        DASH_r3_label,
        DASH_r3_output
    ], layout=CARD_STYLE)

    r4_card = widgets.VBox([
        DASH_r4_label,
        DASH_r4_value
    ], layout=CARD_STYLE)

    report_buttons = widgets.GridBox([
        DASH_r5_button, DASH_r6_button, DASH_r7_button,
        DASH_r8_button, DASH_r9_button, DASH_r10_button
    ], layout={
        'grid_template_columns': 'repeat(auto-fit, minmax(220px, 1fr))',
        'grid_gap': '15px',
        'margin': '20px 0'
    })

    dashboard_layout = widgets.VBox([
        DASH_title,
        SECTION_SPACING,
        widgets.HBox([r1_card, r2_card], layout={'justify_content': 'space-between', 'flex_wrap': 'wrap'}),
        widgets.HBox([r3_card, r4_card], layout={'justify_content': 'space-between', 'flex_wrap': 'wrap'}),
        DASH_download_area,  # Add download area to layout
        SEPARATOR,
        widgets.Label("Detailed Reports:", style=SUBHEADER_STYLE),
        report_buttons,
        SECTION_SPACING,
        DASH_back_button
    ], layout={
        '-with': '30px',
        'max_width': '1400px',
        'margin': '0 auto',
        'background': '#f5f5f5'
    })

    # Add CSS for hover effects
    DASH_r1_button.style._css = {'transition': 'background-color 0.3s'}
    DASH_r1_button.style._hover = {'background_color': '#45a049'}
    for btn in [DASH_r5_button, DASH_r6_button, DASH_r7_button, DASH_r8_button, DASH_r9_button, DASH_r10_button]:
        btn.style._css = {'transition': 'background-color 0.3s'}
        btn.style._hover = {'background_color': '#1976D2'}
    DASH_back_button.style._css = {'transition': 'background-color 0.3s'}
    DASH_back_button.style._hover = {'background_color': '#e0e0e0'}

    display(dashboard_layout)
    refresh_dashboard()

# New Tab: View All Orders
with tb_all.output_to('View All Orders'):
    ORDERS_title = widgets.Label("Orders Received Today")
    ORDERS_title.style = HEADER_STYLE
    ORDERS_output = widgets.Output()
    # Add dedicated download area
    ORDERS_download_area = widgets.Output(layout={'height': '80px', 'min_height': '80px', 'margin': '10px 0', 'overflow': 'hidden'})
    ORDERS_back_button = widgets.Button(
        description="Back to Dashboard",
        layout={'width': '220px', 'margin': '10px 0'},
        tooltip="Return to the manager dashboard"
    )
    style_button(ORDERS_back_button, NEUTRAL_COLOR)

    # Add Excel export button
    ORDERS_excel_standalone_button = widgets.Button(
        description="Export All Data to Excel",
        layout={'width': '220px', 'margin': '10px 0'},
        tooltip="Export all report data to Excel file"
    )
    style_button(ORDERS_excel_standalone_button, SECONDARY_COLOR)

    def refresh_orders_page():
        try:
            c.execute("""
                SELECT O_ID, O_DATE, O_TOTAL_AMOUNT, O_STATUS, MEMBER_M_ID
                FROM ORDERS
                WHERE TO_CHAR(O_DATE, 'YYYY-MM-DD') = TO_CHAR(SYSDATE, 'YYYY-MM-DD')
                ORDER BY O_ID
            """)
            orders = c.fetchall()

            # Create a DataFrame for Excel export
            global orders_df
            orders_df = pd.DataFrame(orders, columns=["Order ID", "Date", "Total Amount",
                                                    "Status", "Member ID"])

            with ORDERS_output:
                clear_output()
                if not orders:
                    display(widgets.Label("No orders found for today."))
                else:
                    page_size = 10
                    pages = [orders[i:i + page_size] for i in range(0, len(orders), page_size)]
                    page_dropdown = widgets.Dropdown(
                        options=[f"Page {i+1}" for i in range(len(pages))],
                        description="Select Page:",
                        style={'description_width': 'initial'},
                        layout={'width': '200px', 'margin': '10px 0'}
                    )
                    next_button = widgets.Button(
                        description="Next Page",
                        layout={'width': '150px', 'margin': '10px 10px'},
                        tooltip="Go to the next page"
                    )
                    ORDERS_excel_button = widgets.Button(
                        description="Export to Excel",
                        layout={'width': '150px', 'margin': '10px 10px'},
                        tooltip="Export report data to Excel file"
                    )
                    style_button(next_button, SECONDARY_COLOR)
                    style_button(ORDERS_excel_standalone_button, SECONDARY_COLOR)

                    def display_page(page_idx):
                        with ORDERS_output:
                            clear_output()
                            display(widgets.Label(f"Page {page_idx + 1} of {len(pages)}", style=SUBHEADER_STYLE))
                            header = widgets.HBox([
                                widgets.Label("Order ID", style=TABLE_HEADER_STYLE, layout={'width': '120px'}),
                                widgets.Label("Date", style=TABLE_HEADER_STYLE, layout={'width': '200px'}),
                                widgets.Label("Total Amount", style={**TABLE_HEADER_STYLE, 'text_align': 'right'}, layout={'width': '150px'}),
                                widgets.Label("Status", style=TABLE_HEADER_STYLE, layout={'width': '150px'}),
                                widgets.Label("Member ID", style=TABLE_HEADER_STYLE, layout={'width': '120px'})
                            ], layout=TABLE_HEADER_LAYOUT)
                            display(header)
                            for i, order in enumerate(pages[page_idx]):
                                bg_color = '#f9f9f9' if i % 2 == 0 else '#ffffff'
                                display(widgets.HBox([
                                    widgets.Label(str(order[0]), style=TABLE_DATA_STYLE, layout={'width': '120px'}),
                                    widgets.Label(str(order[1]), style=TABLE_DATA_STYLE, layout={'width': '200px'}),
                                    widgets.Label(f"${order[2]:,.2f}", style={**TABLE_DATA_STYLE, 'text_align': 'right'}, layout={'width': '150px'}),
                                    widgets.Label(order[3], style=TABLE_DATA_STYLE, layout={'width': '150px'}),
                                    widgets.Label(str(order[4]), style=TABLE_DATA_STYLE, layout={'width': '120px'})
                                ], layout={**TABLE_ROW_LAYOUT, 'background': bg_color}))
                            display(widgets.HBox([page_dropdown, next_button, ORDERS_excel_button]))

                    def on_dropdown_change(change):
                        page_idx = int(change['new'].split()[1]) - 1
                        display_page(page_idx)

                    def on_next_button_clicked(b):
                        current_page = int(page_dropdown.value.split()[1]) - 1
                        if current_page < len(pages) - 1:
                            page_dropdown.value = f"Page {current_page + 2}"
                            display_page(current_page + 1)

                    def on_orders_excel_button_clicked(b):
                        # Use the dedicated download area
                        display_excel_download(
                            ORDERS_download_area,
                            orders_df,
                            "todays_orders.xlsx",
                            "Download Today's Orders Report"
                        )

                    ORDERS_excel_button.on_click(on_orders_excel_button_clicked)
                    page_dropdown.observe(on_dropdown_change, names='value')
                    next_button.on_click(on_next_button_clicked)
                    next_button.disabled = len(pages) <= 1  # Disable if only one page
                    display_page(0)
        except Exception as e:
            with ORDERS_output:
                clear_output()
                display(widgets.Label(f"Error: {str(e)}"))

    def on_orders_excel_standalone_clicked(b):
        try:
            if 'orders_df' in globals():
                # Use the dedicated download area
                display_excel_download(
                    ORDERS_download_area,
                    orders_df,
                    "todays_orders.xlsx",
                    "Download Today's Orders Report"
                )
        except Exception as e:
            with ORDERS_output:
                display(widgets.Label(f"Error generating Excel file: {str(e)}"))

    ORDERS_excel_standalone_button.on_click(on_orders_excel_standalone_clicked)

    def on_orders_back_button_clicked(b):
        # Clear the download area before leaving
        with ORDERS_download_area:
            clear_output()
        showPage("Manager Dashboard")
        refresh_dashboard()
    ORDERS_back_button.on_click(on_orders_back_button_clicked)

    orders_layout = widgets.VBox([
        ORDERS_title,
        SECTION_SPACING,
        ORDERS_output,
        ORDERS_download_area,  # Add download area to layout
        SECTION_SPACING,
        widgets.HBox([ORDERS_back_button, ORDERS_excel_standalone_button])
    ], layout={
        'padding': '30px',
        'max_width': '1400px',
        'margin': '0 auto',
        'background': '#f5f5f5'
    })

    ORDERS_back_button.style._css = {'transition': 'background-color 0.3s'}
    ORDERS_back_button.style._hover = {'background_color': '#e0e0e0'}
    ORDERS_excel_standalone_button.style._css = {'transition': 'background-color 0.3s'}
    ORDERS_excel_standalone_button.style._hover = {'background_color': '#1976D2'}

    display(orders_layout)

# New Tab: Sales Revenue by Date (R5)
with tb_all.output_to('Sales Revenue by Date'):
    R5_title = widgets.Label("Sales Revenue by Date")
    R5_title.style = HEADER_STYLE
    R5_output = widgets.Output()
    # Add a dedicated download area with fixed height
    R5_download_area = widgets.Output(layout={'height': '80px', 'min_height': '80px', 'margin': '10px 0', 'overflow': 'hidden'})
    R5_back_button = widgets.Button(
        description="Back to Dashboard",
        layout={'width': '220px', 'margin': '10px 0'},
        tooltip="Return to the manager dashboard"
    )
    style_button(R5_back_button, NEUTRAL_COLOR)

    # Add Excel export button
    R5_excel_standalone_button = widgets.Button(
        description="Export All Data to Excel",
        layout={'width': '220px', 'margin': '10px 0'},
        tooltip="Export all report data to Excel file"
    )
    style_button(R5_excel_standalone_button, SECONDARY_COLOR)

    def refresh_r5_page():
        try:
            c.execute("""
                SELECT TO_CHAR(O_DATE, 'YYYY-MM-DD'), SUM(O_TOTAL_AMOUNT)
                FROM ORDERS
                GROUP BY TO_CHAR(O_DATE, 'YYYY-MM-DD')
                ORDER BY TO_CHAR(O_DATE, 'YYYY-MM-DD') DESC
            """)
            rows = c.fetchall()

            # Create a DataFrame for Excel export
            global r5_df
            r5_df = pd.DataFrame(rows, columns=["Date", "Total Revenue"])

            with R5_output:
                clear_output()
                if not rows:
                    display(widgets.Label("No sales data available."))
                else:
                    page_size = 10
                    pages = [rows[i:i + page_size] for i in range(0, len(rows), page_size)]
                    page_dropdown = widgets.Dropdown(
                        options=[f"Page {i+1}" for i in range(len(pages))],
                        description="Select Page:",
                        style={'description_width': 'initial'},
                        layout={'width': '200px', 'margin': '10px 0'}
                    )
                    next_button = widgets.Button(
                        description="Next Page",
                        layout={'width': '150px', 'margin': '10px 10px'},
                        tooltip="Go to the next page"
                    )
                    R5_excel_button = widgets.Button(
                        description="Export to Excel",
                        layout={'width': '150px', 'margin': '10px 10px'},
                        tooltip="Export report data to Excel file"
                    )
                    style_button(next_button, SECONDARY_COLOR)
                    style_button(R5_excel_standalone_button, SECONDARY_COLOR)

                    def display_page(page_idx):
                        with R5_output:
                            clear_output()
                            display(widgets.Label(f"Page {page_idx + 1} of {len(pages)}", style=SUBHEADER_STYLE))
                            header = widgets.HBox([
                                widgets.Label("Date", style=TABLE_HEADER_STYLE, layout={'width': '200px'}),
                                widgets.Label("Total Revenue", style={**TABLE_HEADER_STYLE, 'text_align': 'right'}, layout={'width': '200px'})
                            ], layout=TABLE_HEADER_LAYOUT)
                            display(header)
                            for i, row in enumerate(pages[page_idx]):
                                bg_color = '#f9f9f9' if i % 2 == 0 else '#ffffff'
                                display(widgets.HBox([
                                    widgets.Label(row[0], style=TABLE_DATA_STYLE, layout={'width': '200px'}),
                                    widgets.Label(f"${row[1]:,.2f}", style={**TABLE_DATA_STYLE, 'text_align': 'right'}, layout={'width': '200px'})
                                ], layout={**TABLE_ROW_LAYOUT, 'background': bg_color}))
                            display(widgets.HBox([page_dropdown, next_button, R5_excel_button]))

                    def on_dropdown_change(change):
                        page_idx = int(change['new'].split()[1]) - 1
                        display_page(page_idx)

                    def on_next_button_clicked(b):
                        current_page = int(page_dropdown.value.split()[1]) - 1
                        if current_page < len(pages) - 1:
                            page_dropdown.value = f"Page {current_page + 2}"
                            display_page(current_page + 1)

                    def on_r5_excel_button_clicked(b):
                        # Use the dedicated download area
                        display_excel_download(
                            R5_download_area,
                            r5_df,
                            "sales_revenue_by_date.xlsx",
                            "Download Sales Revenue Report"
                        )

                    R5_excel_button.on_click(on_r5_excel_button_clicked)
                    page_dropdown.observe(on_dropdown_change, names='value')
                    next_button.on_click(on_next_button_clicked)
                    next_button.disabled = len(pages) <= 1
                    display_page(0)
        except Exception as e:
            with R5_output:
                clear_output()
                display(widgets.Label(f"Error: {str(e)}"))

    def on_r5_excel_standalone_clicked(b):
        try:
            if 'r5_df' in globals():
                # Use the dedicated download area
                display_excel_download(
                    R5_download_area,
                    r5_df,
                    "sales_revenue_by_date.xlsx",
                    "Download Sales Revenue Report"
                )
        except Exception as e:
            with R5_output:
                display(widgets.Label(f"Error generating Excel file: {str(e)}"))

    R5_excel_standalone_button.on_click(on_r5_excel_standalone_clicked)

    def on_r5_back_button_clicked(b):
        # Clear the download area before leaving
        with R5_download_area:
            clear_output()
        showPage("Manager Dashboard")
        refresh_dashboard()
    R5_back_button.on_click(on_r5_back_button_clicked)

    r5_layout = widgets.VBox([
        R5_title,
        SECTION_SPACING,
        R5_output,
        R5_download_area,  # Add download area to layout
        SECTION_SPACING,
        widgets.HBox([R5_back_button, R5_excel_standalone_button])
    ], layout={
        'padding': '30px',
        'max_width': '1400px',
        'margin': '0 auto',
        'background': '#f5f5f5'
    })

    R5_back_button.style._css = {'transition': 'background-color 0.3s'}
    R5_back_button.style._hover = {'background_color': '#e0e0e0'}
    R5_excel_standalone_button.style._css = {'transition': 'background-color 0.3s'}
    R5_excel_standalone_button.style._hover = {'background_color': '#1976D2'}

    display(r5_layout)

# New Tab: Sales Quantity by Date (R6)
with tb_all.output_to('Sales Quantity by Date'):
    R6_title = widgets.Label("Sales Quantity by Date")
    R6_title.style = HEADER_STYLE
    R6_output = widgets.Output()
    # Add dedicated download area
    R6_download_area = widgets.Output(layout={'height': '80px', 'min_height': '80px', 'margin': '10px 0', 'overflow': 'hidden'})
    R6_back_button = widgets.Button(
        description="Back to Dashboard",
        layout={'width': '220px', 'margin': '10px 0'},
        tooltip="Return to the manager dashboard"
    )
    style_button(R6_back_button, NEUTRAL_COLOR)

    # Add Excel export button
    R6_excel_standalone_button = widgets.Button(
        description="Export All Data to Excel",
        layout={'width': '220px', 'margin': '10px 0'},
        tooltip="Export all report data to Excel file"
    )
    style_button(R6_excel_standalone_button, SECONDARY_COLOR)

    def refresh_r6_page():
        try:
            c.execute("""
                SELECT TO_CHAR(O.O_DATE, 'YYYY-MM-DD'), SUM(OL.OL_QTY)
                FROM ORDERS O
                INNER JOIN ORDER_LINE OL ON O.O_ID = OL.ORDERS_O_ID
                GROUP BY TO_CHAR(O.O_DATE, 'YYYY-MM-DD')
                ORDER BY TO_CHAR(O.O_DATE, 'YYYY-MM-DD') DESC
            """)
            rows = c.fetchall()

            # Create a DataFrame for Excel export
            global r6_df
            r6_df = pd.DataFrame(rows, columns=["Date", "Total Quantity"])

            with R6_output:
                clear_output()
                if not rows:
                    display(widgets.Label("No sales data available."))
                else:
                    page_size = 10
                    pages = [rows[i:i + page_size] for i in range(0, len(rows), page_size)]
                    page_dropdown = widgets.Dropdown(
                        options=[f"Page {i+1}" for i in range(len(pages))],
                        description="Select Page:",
                        style={'description_width': 'initial'},
                        layout={'width': '200px', 'margin': '10px 0'}
                    )
                    next_button = widgets.Button(
                        description="Next Page",
                        layout={'width': '150px', 'margin': '10px 10px'},
                        tooltip="Go to the next page"
                    )
                    R6_excel_button = widgets.Button(
                        description="Export to Excel",
                        layout={'width': '150px', 'margin': '10px 10px'},
                        tooltip="Export report data to Excel file"
                    )
                    style_button(next_button, SECONDARY_COLOR)
                    style_button(R6_excel_standalone_button, SECONDARY_COLOR)

                    def display_page(page_idx):
                        with R6_output:
                            clear_output()
                            display(widgets.Label(f"Page {page_idx + 1} of {len(pages)}", style=SUBHEADER_STYLE))
                            header = widgets.HBox([
                                widgets.Label("Date", style=TABLE_HEADER_STYLE, layout={'width': '200px'}),
                                widgets.Label("Total Quantity", style={**TABLE_HEADER_STYLE, 'text_align': 'right'}, layout={'width': '200px'})
                            ], layout=TABLE_HEADER_LAYOUT)
                            display(header)
                            for i, row in enumerate(pages[page_idx]):
                                bg_color = '#f9f9f9' if i % 2 == 0 else '#ffffff'
                                display(widgets.HBox([
                                    widgets.Label(row[0], style=TABLE_DATA_STYLE, layout={'width': '200px'}),
                                    widgets.Label(f"{row[1]:,} pcs", style={**TABLE_DATA_STYLE, 'text_align': 'right'}, layout={'width': '200px'})
                                ], layout={**TABLE_ROW_LAYOUT, 'background': bg_color}))
                            display(widgets.HBox([page_dropdown, next_button, R6_excel_button]))

                    def on_dropdown_change(change):
                        page_idx = int(change['new'].split()[1]) - 1
                        display_page(page_idx)

                    def on_next_button_clicked(b):
                        current_page = int(page_dropdown.value.split()[1]) - 1
                        if current_page < len(pages) - 1:
                            page_dropdown.value = f"Page {current_page + 2}"
                            display_page(current_page + 1)

                    def on_r6_excel_button_clicked(b):
                        # Use the dedicated download area
                        display_excel_download(
                            R6_download_area,
                            r6_df,
                            "sales_quantity_by_date.xlsx",
                            "Download Sales Quantity Report"
                        )

                    R6_excel_button.on_click(on_r6_excel_button_clicked)
                    page_dropdown.observe(on_dropdown_change, names='value')
                    next_button.on_click(on_next_button_clicked)
                    next_button.disabled = len(pages) <= 1
                    display_page(0)
        except Exception as e:
            with R6_output:
                clear_output()
                display(widgets.Label(f"Error: {str(e)}"))

    def on_r6_excel_standalone_clicked(b):
        try:
            if 'r6_df' in globals():
                # Use the dedicated download area
                display_excel_download(
                    R6_download_area,
                    r6_df,
                    "sales_quantity_by_date.xlsx",
                    "Download Sales Quantity Report"
                )
        except Exception as e:
            with R6_output:
                display(widgets.Label(f"Error generating Excel file: {str(e)}"))

    R6_excel_standalone_button.on_click(on_r6_excel_standalone_clicked)

    def on_r6_back_button_clicked(b):
        # Clear the download area before leaving
        with R6_download_area:
            clear_output()
        showPage("Manager Dashboard")
        refresh_dashboard()
    R6_back_button.on_click(on_r6_back_button_clicked)

    r6_layout = widgets.VBox([
        R6_title,
        SECTION_SPACING,
        R6_output,
        R6_download_area,  # Add download area to layout that was missing
        SECTION_SPACING,
        widgets.HBox([R6_back_button, R6_excel_standalone_button])
    ], layout={
        'padding': '30px',
        'max_width': '1400px',
        'margin': '0 auto',
        'background': '#f5f5f5'
    })

    R6_back_button.style._css = {'transition': 'background-color 0.3s'}
    R6_back_button.style._hover = {'background_color': '#e0e0e0'}
    R6_excel_standalone_button.style._css = {'transition': 'background-color 0.3s'}
    R6_excel_standalone_button.style._hover = {'background_color': '#1976D2'}

    display(r6_layout)

# New Tab: Trade-in Discounts by Date (R7)
with tb_all.output_to('Trade-in Discounts by Date'):
    R7_title = widgets.Label("Trade-in Discounts by Date")
    R7_title.style = HEADER_STYLE
    R7_output = widgets.Output()
    # Add dedicated download area
    R7_download_area = widgets.Output(layout={'height': '80px', 'min_height': '80px', 'margin': '10px 0', 'overflow': 'hidden'})
    R7_back_button = widgets.Button(
        description="Back to Dashboard",
        layout={'width': '220px', 'margin': '10px 0'},
        tooltip="Return to the manager dashboard"
    )
    style_button(R7_back_button, NEUTRAL_COLOR)

    # Add Excel export button
    R7_excel_standalone_button = widgets.Button(
        description="Export All Data to Excel",
        layout={'width': '220px', 'margin': '10px 0'},
        tooltip="Export all report data to Excel file"
    )
    style_button(R7_excel_standalone_button, SECONDARY_COLOR)

    def refresh_r7_page():
        try:
            c.execute("""
                SELECT TO_CHAR(O.O_DATE, 'YYYY-MM-DD'), SUM(T.TRADE_IN_PRICE)
                FROM ORDERS O
                LEFT JOIN TRADE_IN_DEVICE T ON O.O_ID = T.ORDERS_O_ID
                GROUP BY TO_CHAR(O.O_DATE, 'YYYY-MM-DD')
                ORDER BY TO_CHAR(O.O_DATE, 'YYYY-MM-DD') DESC
            """)
            rows = c.fetchall()

            # Create a DataFrame for Excel export
            global r7_df
            r7_df = pd.DataFrame(rows, columns=["Date", "Total Trade-in Discount"])

            with R7_output:
                clear_output()
                if not rows:
                    display(widgets.Label("No trade-in data available."))
                else:
                    page_size = 10
                    pages = [rows[i:i + page_size] for i in range(0, len(rows), page_size)]
                    page_dropdown = widgets.Dropdown(
                        options=[f"Page {i+1}" for i in range(len(pages))],
                        description="Select Page:",
                        style={'description_width': 'initial'},
                        layout={'width': '200px', 'margin': '10px 0'}
                    )
                    next_button = widgets.Button(
                        description="Next Page",
                        layout={'width': '150px', 'margin': '10px 10px'},
                        tooltip="Go to the next page"
                    )
                    R7_excel_button = widgets.Button(
                        description="Export to Excel",
                        layout={'width': '150px', 'margin': '10px 10px'},
                        tooltip="Export report data to Excel file"
                    )
                    style_button(next_button, SECONDARY_COLOR)
                    style_button(R7_excel_standalone_button, SECONDARY_COLOR)

                    def display_page(page_idx):
                        with R7_output:
                            clear_output()
                            display(widgets.Label(f"Page {page_idx + 1} of {len(pages)}", style=SUBHEADER_STYLE))
                            header = widgets.HBox([
                                widgets.Label("Date", style=TABLE_HEADER_STYLE, layout={'width': '200px'}),
                                widgets.Label("Total Discount", style={**TABLE_HEADER_STYLE, 'text_align': 'right'}, layout={'width': '200px'})
                            ], layout=TABLE_HEADER_LAYOUT)
                            display(header)
                            for i, row in enumerate(pages[page_idx]):
                                bg_color = '#f9f9f9' if i % 2 == 0 else '#ffffff'
                                display(widgets.HBox([
                                    widgets.Label(row[0], style=TABLE_DATA_STYLE, layout={'width': '200px'}),
                                    widgets.Label(f"${row[1] or 0:,.2f}", style={**TABLE_DATA_STYLE, 'text_align': 'right'}, layout={'width': '200px'})
                                ], layout={**TABLE_ROW_LAYOUT, 'background': bg_color}))
                            display(widgets.HBox([page_dropdown, next_button, R7_excel_button]))

                    def on_dropdown_change(change):
                        page_idx = int(change['new'].split()[1]) - 1
                        display_page(page_idx)

                    def on_next_button_clicked(b):
                        current_page = int(page_dropdown.value.split()[1]) - 1
                        if current_page < len(pages) - 1:
                            page_dropdown.value = f"Page {current_page + 2}"
                            display_page(current_page + 1)

                    def on_r7_excel_button_clicked(b):
                        # Use the dedicated download area
                        display_excel_download(
                            R7_download_area,
                            r7_df,
                            "trade_in_discounts_by_date.xlsx",
                            "Download Trade-in Discounts Report"
                        )

                    R7_excel_button.on_click(on_r7_excel_button_clicked)
                    page_dropdown.observe(on_dropdown_change, names='value')
                    next_button.on_click(on_next_button_clicked)
                    next_button.disabled = len(pages) <= 1
                    display_page(0)
        except Exception as e:
            with R7_output:
                clear_output()
                display(widgets.Label(f"Error: {str(e)}"))

    def on_r7_excel_standalone_clicked(b):
        try:
            if 'r7_df' in globals():
                # Use the dedicated download area
                display_excel_download(
                    R7_download_area,
                    r7_df,
                    "trade_in_discounts_by_date.xlsx",
                    "Download Trade-in Discounts Report"
                )
        except Exception as e:
            with R7_output:
                display(widgets.Label(f"Error generating Excel file: {str(e)}"))

    R7_excel_standalone_button.on_click(on_r7_excel_standalone_clicked)

    def on_r7_back_button_clicked(b):
        # Clear the download area before leaving
        with R7_download_area:
            clear_output()
        showPage("Manager Dashboard")
        refresh_dashboard()
    R7_back_button.on_click(on_r7_back_button_clicked)

    r7_layout = widgets.VBox([
        R7_title,
        SECTION_SPACING,
        R7_output,
        R7_download_area,  # Add download area to layout
        SECTION_SPACING,
        widgets.HBox([R7_back_button, R7_excel_standalone_button])
    ], layout={
        'padding': '30px',
        'max_width': '1400px',
        'margin': '0 auto',
        'background': '#f5f5f5'
    })

    R7_back_button.style._css = {'transition': 'background-color 0.3s'}
    R7_back_button.style._hover = {'background_color': '#e0e0e0'}
    R7_excel_standalone_button.style._css = {'transition': 'background-color 0.3s'}
    R7_excel_standalone_button.style._hover = {'background_color': '#1976D2'}

    display(r7_layout)

# New Tab: Old Devices Received (R8)
with tb_all.output_to('Old Devices Received'):
    R8_title = widgets.Label("Old Devices Received (Accepted)")
    R8_title.style = HEADER_STYLE
    R8_output = widgets.Output()
    # Add dedicated download area
    R8_download_area = widgets.Output(layout={'height': '80px', 'min_height': '80px', 'margin': '10px 0', 'overflow': 'hidden'})
    R8_back_button = widgets.Button(
        description="Back to Dashboard",
        layout={'width': '220px', 'margin': '10px 0'},
        tooltip="Return to the manager dashboard"
    )
    style_button(R8_back_button, NEUTRAL_COLOR)

    # Add Excel export button
    R8_excel_standalone_button = widgets.Button(
        description="Export All Data to Excel",
        layout={'width': '220px', 'margin': '10px 0'},
        tooltip="Export all report data to Excel file"
    )
    style_button(R8_excel_standalone_button, SECONDARY_COLOR)

    def refresh_r8_page():
        try:
            c.execute("""
                SELECT T.T_ID, T.SERIAL_NUMBER, T.CONDITION, T.TRADE_IN_PRICE, T.T_STATUS, T.ORDERS_O_ID
                FROM TRADE_IN_DEVICE T
                WHERE T.T_STATUS = 'Accepted'
                ORDER BY T.T_ID
            """)
            rows = c.fetchall()

            # Create a DataFrame for Excel export
            global r8_df
            r8_df = pd.DataFrame(rows, columns=["Trade-in ID", "Serial Number", "Condition",
                                               "Trade-in Price", "Status", "Order ID"])

            with R8_output:
                clear_output()
                if not rows:
                    display(widgets.Label("No accepted trade-in devices found."))
                else:
                    page_size = 10
                    pages = [rows[i:i + page_size] for i in range(0, len(rows), page_size)]
                    page_dropdown = widgets.Dropdown(
                        options=[f"Page {i+1}" for i in range(len(pages))],
                        description="Select Page:",
                        style={'description_width': 'initial'},
                        layout={'width': '200px', 'margin': '10px 0'}
                    )
                    next_button = widgets.Button(
                        description="Next Page",
                        layout={'width': '150px', 'margin': '10px 10px'},
                        tooltip="Go to the next page"
                    )
                    R8_excel_button = widgets.Button(
                        description="Export to Excel",
                        layout={'width': '150px', 'margin': '10px 10px'},
                        tooltip="Export report data to Excel file"
                    )
                    style_button(next_button, SECONDARY_COLOR)
                    style_button(R8_excel_standalone_button, SECONDARY_COLOR)

                    def display_page(page_idx):
                        with R8_output:
                            clear_output()
                            display(widgets.Label(f"Page {page_idx + 1} of {len(pages)}", style=SUBHEADER_STYLE))
                            header = widgets.HBox([
                                widgets.Label("Trade-in ID", style=TABLE_HEADER_STYLE, layout={'width': '120px'}),
                                widgets.Label("Serial Number", style=TABLE_HEADER_STYLE, layout={'width': '200px'}),
                                widgets.Label("Condition", style=TABLE_HEADER_STYLE, layout={'width': '150px'}),
                                widgets.Label("Trade-in Price", style={**TABLE_HEADER_STYLE, 'text_align': 'right'}, layout={'width': '150px'}),
                                widgets.Label("Status", style=TABLE_HEADER_STYLE, layout={'width': '150px'}),
                                widgets.Label("Order ID", style=TABLE_HEADER_STYLE, layout={'width': '120px'})
                            ], layout=TABLE_HEADER_LAYOUT)
                            display(header)
                            for i, row in enumerate(pages[page_idx]):
                                bg_color = '#f9f9f9' if i % 2 == 0 else '#ffffff'
                                display(widgets.HBox([
                                    widgets.Label(str(row[0]), style=TABLE_DATA_STYLE, layout={'width': '120px'}),
                                    widgets.Label(str(row[1]), style=TABLE_DATA_STYLE, layout={'width': '200px'}),
                                    widgets.Label(str(row[2]), style=TABLE_DATA_STYLE, layout={'width': '150px'}),
                                    widgets.Label(f"${row[3]:,.2f}" if row[3] is not None else "$0.00", style={**TABLE_DATA_STYLE, 'text_align': 'right'}, layout={'width': '150px'}),
                                    widgets.Label(str(row[4]), style=TABLE_DATA_STYLE, layout={'width': '150px'}),
                                    widgets.Label(str(row[5]), style=TABLE_DATA_STYLE, layout={'width': '120px'})
                                ], layout={**TABLE_ROW_LAYOUT, 'background': bg_color}))
                            display(widgets.HBox([page_dropdown, next_button, R8_excel_button]))

                    def on_dropdown_change(change):
                        page_idx = int(change['new'].split()[1]) - 1
                        display_page(page_idx)

                    def on_next_button_clicked(b):
                        current_page = int(page_dropdown.value.split()[1]) - 1
                        if current_page < len(pages) - 1:
                            page_dropdown.value = f"Page {current_page + 2}"
                            display_page(current_page + 1)

                    def on_r8_excel_button_clicked(b):
                        # Use the dedicated download area
                        display_excel_download(
                            R8_download_area,
                            r8_df,
                            "old_devices_received.xlsx",
                            "Download Old Devices Report"
                        )

                    R8_excel_button.on_click(on_r8_excel_button_clicked)
                    page_dropdown.observe(on_dropdown_change, names='value')
                    next_button.on_click(on_next_button_clicked)
                    next_button.disabled = len(pages) <= 1
                    display_page(0)
        except Exception as e:
            with R8_output:
                clear_output()
                display(widgets.Label(f"Error: {str(e)}"))

    def on_r8_excel_standalone_clicked(b):
        try:
            if 'r8_df' in globals():
                # Use the dedicated download area
                display_excel_download(
                    R8_download_area,
                    r8_df,
                    "old_devices_received.xlsx",
                    "Download Old Devices Report"
                )
        except Exception as e:
            with R8_output:
                display(widgets.Label(f"Error generating Excel file: {str(e)}"))

    R8_excel_standalone_button.on_click(on_r8_excel_standalone_clicked)

    def on_r8_back_button_clicked(b):
        # Clear the download area before leaving
        with R8_download_area:
            clear_output()
        showPage("Manager Dashboard")
        refresh_dashboard()
    R8_back_button.on_click(on_r8_back_button_clicked)

    r8_layout = widgets.VBox([
        R8_title,
        SECTION_SPACING,
        R8_output,
        R8_download_area,  # Add download area to layout
        SECTION_SPACING,
        widgets.HBox([R8_back_button, R8_excel_standalone_button])
    ], layout={
        'padding': '30px',
        'max_width': '1400px',
        'margin': '0 auto',
        'background': '#f5f5f5'
    })

    R8_back_button.style._css = {'transition': 'background-color 0.3s'}
    R8_back_button.style._hover = {'background_color': '#e0e0e0'}
    R8_excel_standalone_button.style._css = {'transition': 'background-color 0.3s'}
    R8_excel_standalone_button.style._hover = {'background_color': '#1976D2'}

    display(r8_layout)

# New Tab: Out-of-Stock Products (R9)
with tb_all.output_to('Out-of-Stock Products'):
    R9_title = widgets.Label("Out-of-Stock Products")
    R9_title.style = HEADER_STYLE
    R9_output = widgets.Output()
    # Add dedicated download area
    R9_download_area = widgets.Output(layout={'height': '80px', 'min_height': '80px', 'margin': '10px 0', 'overflow': 'hidden'})
    R9_back_button = widgets.Button(
        description="Back to Dashboard",
        layout={'width': '220px', 'margin': '10px 0'},
        tooltip="Return to the manager dashboard"
    )
    style_button(R9_back_button, NEUTRAL_COLOR)

    # Add Excel export button
    R9_excel_standalone_button = widgets.Button(
        description="Export All Data to Excel",
        layout={'width': '220px', 'margin': '10px 0'},
        tooltip="Export all report data to Excel file"
    )
    style_button(R9_excel_standalone_button, SECONDARY_COLOR)

    def refresh_r9_page():
        try:
            c.execute("""
                SELECT P_ID, P_NAME, P_DESCRIPTION, P_PRICE, P_CAPACITY, P_MANUFACTURER
                FROM PRODUCT
                WHERE P_INVENTORY = 0
                ORDER BY P_ID
            """)
            rows = c.fetchall()

            # Create a DataFrame for Excel export
            global r9_df
            r9_df = pd.DataFrame(rows, columns=["Product ID", "Name", "Description",
                                               "Price", "Capacity", "Manufacturer"])

            with R9_output:
                clear_output()
                if not rows:
                    display(widgets.Label("No out-of-stock products found."))
                else:
                    page_size = 10
                    pages = [rows[i:i + page_size] for i in range(0, len(rows), page_size)]
                    page_dropdown = widgets.Dropdown(
                        options=[f"Page {i+1}" for i in range(len(pages))],
                        description="Select Page:",
                        style={'description_width': 'initial'},
                        layout={'width': '200px', 'margin': '10px 0'}
                    )
                    next_button = widgets.Button(
                        description="Next Page",
                        layout={'width': '150px', 'margin': '10px 10px'},
                        tooltip="Go to the next page"
                    )
                    R9_excel_button = widgets.Button(
                        description="Export to Excel",
                        layout={'width': '150px', 'margin': '10px 10px'},
                        tooltip="Export report data to Excel file"
                    )
                    style_button(next_button, SECONDARY_COLOR)
                    style_button(R9_excel_standalone_button, SECONDARY_COLOR)

                    def display_page(page_idx):
                        with R9_output:
                            clear_output()
                            display(widgets.Label(f"Page {page_idx + 1} of {len(pages)}", style=SUBHEADER_STYLE))
                            header = widgets.HBox([
                                widgets.Label("ID", style=TABLE_HEADER_STYLE, layout={'width': '80px'}),
                                widgets.Label("Name", style=TABLE_HEADER_STYLE, layout={'width': '200px'}),
                                widgets.Label("Description", style=TABLE_HEADER_STYLE, layout={'width': '250px'}),
                                widgets.Label("Price", style={**TABLE_HEADER_STYLE, 'text_align': 'right'}, layout={'width': '100px'}),
                                widgets.Label("Capacity", style=TABLE_HEADER_STYLE, layout={'width': '100px'}),
                                widgets.Label("Manufacturer", style=TABLE_HEADER_STYLE, layout={'width': '150px'})
                            ], layout=TABLE_HEADER_LAYOUT)
                            display(header)
                            for i, row in enumerate(pages[page_idx]):
                                bg_color = '#f9f9f9' if i % 2 == 0 else '#ffffff'

                                # Handle price formatting - convert string to float first
                                try:
                                    price = float(row[3])
                                    price_formatted = f"${price:,.2f}"
                                except (ValueError, TypeError):
                                    # If conversion fails, just display the price as is
                                    price_formatted = f"${row[3]}" if row[3] else "$0.00"

                                display(widgets.HBox([
                                    widgets.Label(str(row[0]), style=TABLE_DATA_STYLE, layout={'width': '80px'}),
                                    widgets.Label(row[1], style=TABLE_DATA_STYLE, layout={'width': '200px'}),
                                    widgets.Label(row[2], style=TABLE_DATA_STYLE, layout={'width': '250px'}),
                                    widgets.Label(price_formatted, style={**TABLE_DATA_STYLE, 'text_align': 'right'}, layout={'width': '100px'}),
                                    widgets.Label(row[4], style=TABLE_DATA_STYLE, layout={'width': '100px'}),
                                    widgets.Label(row[5], style=TABLE_DATA_STYLE, layout={'width': '150px'})
                                ], layout={**TABLE_ROW_LAYOUT, 'background': bg_color}))
                            display(widgets.Label(f"Total out-of-stock products: {len(pages[page_idx])}", style={'font_size': '16px', 'margin': '10px 0'}))
                            display(widgets.HBox([page_dropdown, next_button, R9_excel_button]))

                    def on_dropdown_change(change):
                        page_idx = int(change['new'].split()[1]) - 1
                        display_page(page_idx)

                    def on_next_button_clicked(b):
                        current_page = int(page_dropdown.value.split()[1]) - 1
                        if current_page < len(pages) - 1:
                            page_dropdown.value = f"Page {current_page + 2}"
                            display_page(current_page + 1)

                    def on_r9_excel_button_clicked(b):
                        # Use the dedicated download area
                        display_excel_download(
                            R9_download_area,
                            r9_df,
                            "out_of_stock_products.xlsx",
                            "Download Out-of-Stock Products Report"
                        )

                    R9_excel_button.on_click(on_r9_excel_button_clicked)
                    page_dropdown.observe(on_dropdown_change, names='value')
                    next_button.on_click(on_next_button_clicked)
                    next_button.disabled = len(pages) <= 1
                    display_page(0)
        except Exception as e:
            with R9_output:
                clear_output()
                display(widgets.Label(f"Error: {str(e)}"))

    def on_r9_excel_standalone_clicked(b):
        try:
            if 'r9_df' in globals():
                # Use the dedicated download area
                display_excel_download(
                    R9_download_area,
                    r9_df,
                    "out_of_stock_products.xlsx",
                    "Download Out-of-Stock Products Report"
                )
        except Exception as e:
            with R9_output:
                display(widgets.Label(f"Error generating Excel file: {str(e)}"))

    R9_excel_standalone_button.on_click(on_r9_excel_standalone_clicked)

    def on_r9_back_button_clicked(b):
        # Clear the download area before leaving
        with R9_download_area:
            clear_output()
        showPage("Manager Dashboard")
        refresh_dashboard()
    R9_back_button.on_click(on_r9_back_button_clicked)

    r9_layout = widgets.VBox([
        R9_title,
        SECTION_SPACING,
        R9_output,
        R9_download_area,  # Add download area to layout
        SECTION_SPACING,
        widgets.HBox([R9_back_button, R9_excel_standalone_button])
    ], layout={
        'padding': '30px',
        'max_width': '1400px',
        'margin': '0 auto',
        'background': '#f5f5f5'
    })

    R9_back_button.style._css = {'transition': 'background-color 0.3s'}
    R9_back_button.style._hover = {'background_color': '#e0e0e0'}
    R9_excel_standalone_button.style._css = {'transition': 'background-color 0.3s'}
    R9_excel_standalone_button.style._hover = {'background_color': '#1976D2'}

    display(r9_layout)

# New Tab: Top 10 Customers (R10)
with tb_all.output_to('Top 10 Customers'):
    R10_title = widgets.Label("Top 10 Customers by Number of Sales")
    R10_title.style = HEADER_STYLE
    R10_output = widgets.Output()
    # Add dedicated download area
    R10_download_area = widgets.Output(layout={'height': '80px', 'min_height': '80px', 'margin': '10px 0', 'overflow': 'hidden'})
    R10_back_button = widgets.Button(
        description="Back to Dashboard",
        layout={'width': '220px', 'margin': '10px 0'},
        tooltip="Return to the manager dashboard"
    )
    style_button(R10_back_button, NEUTRAL_COLOR)

    # Add Excel export button
    R10_excel_standalone_button = widgets.Button(
        description="Export All Data to Excel",
        layout={'width': '220px', 'margin': '10px 0'},
        tooltip="Export all report data to Excel file"
    )
    style_button(R10_excel_standalone_button, SECONDARY_COLOR)

    def refresh_r10_page():
        try:
            c.execute("""
                SELECT M.M_ID, M.FNAME || ' ' || M.LNAME, COUNT(*)
                FROM MEMBER M
                INNER JOIN ORDERS O ON M.M_ID = O.MEMBER_M_ID
                GROUP BY M.M_ID, M.FNAME, M.LNAME
                ORDER BY COUNT(*) DESC
                FETCH FIRST 10 ROWS ONLY
            """)
            rows = c.fetchall()

            # Create a DataFrame for Excel export
            global r10_df
            r10_df = pd.DataFrame(rows, columns=["Member ID", "Customer Name", "Order Count"])

            with R10_output:
                clear_output()
                if not rows:
                    display(widgets.Label("No customer sales data available."))
                else:
                    page_size = 10
                    pages = [rows[i:i + page_size] for i in range(0, len(rows), page_size)]
                    page_dropdown = widgets.Dropdown(
                        options=[f"Page {i+1}" for i in range(len(pages))],
                        description="Select Page:",
                        style={'description_width': 'initial'},
                        layout={'width': '200px', 'margin': '10px 0'}
                    )
                    next_button = widgets.Button(
                        description="Next Page",
                        layout={'width': '150px', 'margin': '10px 10px'},
                        tooltip="Go to the next page"
                    )
                    R10_excel_button = widgets.Button(
                        description="Export to Excel",
                        layout={'width': '150px', 'margin': '10px 10px'},
                        tooltip="Export report data to Excel file"
                    )
                    style_button(next_button, SECONDARY_COLOR)
                    style_button(R10_excel_standalone_button, SECONDARY_COLOR)

                    def display_page(page_idx):
                        with R10_output:
                            clear_output()
                            display(widgets.Label(f"Page {page_idx + 1} of {len(pages)}", style=SUBHEADER_STYLE))
                            header = widgets.HBox([
                                widgets.Label("Member ID", style=TABLE_HEADER_STYLE, layout={'width': '150px'}),
                                widgets.Label("Name", style=TABLE_HEADER_STYLE, layout={'width': '350px'}),
                                widgets.Label("Order Count", style={**TABLE_HEADER_STYLE, 'text_align': 'right'}, layout={'width': '150px'})
                            ], layout=TABLE_HEADER_LAYOUT)
                            display(header)
                            for i, row in enumerate(pages[page_idx]):
                                bg_color = '#f9f9f9' if i % 2 == 0 else '#ffffff'
                                display(widgets.HBox([
                                    widgets.Label(str(row[0]), style=TABLE_DATA_STYLE, layout={'width': '150px'}),
                                    widgets.Label(row[1], style=TABLE_DATA_STYLE, layout={'width': '350px'}),
                                    widgets.Label(str(row[2]), style={**TABLE_DATA_STYLE, 'text_align': 'right'}, layout={'width': '150px'})
                                ], layout={**TABLE_ROW_LAYOUT, 'background': bg_color}))
                            display(widgets.HBox([page_dropdown, next_button, R10_excel_button]))

                    def on_dropdown_change(change):
                        page_idx = int(change['new'].split()[1]) - 1
                        display_page(page_idx)

                    def on_next_button_clicked(b):
                        current_page = int(page_dropdown.value.split()[1]) - 1
                        if current_page < len(pages) - 1:
                            page_dropdown.value = f"Page {current_page + 2}"
                            display_page(current_page + 1)

                    def on_r10_excel_button_clicked(b):
                        # Use the dedicated download area
                        display_excel_download(
                            R10_download_area,
                            r10_df,
                            "top_10_customers.xlsx",
                            "Download Top Customers Report"
                        )

                    R10_excel_button.on_click(on_r10_excel_button_clicked)
                    page_dropdown.observe(on_dropdown_change, names='value')
                    next_button.on_click(on_next_button_clicked)
                    next_button.disabled = len(pages) <= 1
                    display_page(0)
        except Exception as e:
            with R10_output:
                clear_output()
                display(widgets.Label(f"Error: {str(e)}"))

    def on_r10_excel_standalone_clicked(b):
        try:
            if 'r10_df' in globals():
                # Use the dedicated download area
                display_excel_download(
                    R10_download_area,
                    r10_df,
                    "top_10_customers.xlsx",
                    "Download Top Customers Report"
                )
        except Exception as e:
            with R10_output:
                display(widgets.Label(f"Error generating Excel file: {str(e)}"))

    R10_excel_standalone_button.on_click(on_r10_excel_standalone_clicked)

    def on_r10_back_button_clicked(b):
        # Clear the download area before leaving
        with R10_download_area:
            clear_output()
        showPage("Manager Dashboard")
        refresh_dashboard()
    R10_back_button.on_click(on_r10_back_button_clicked)

    r10_layout = widgets.VBox([
        R10_title,
        SECTION_SPACING,
        R10_output,
        R10_download_area,  # Add download area to layout
        SECTION_SPACING,
        widgets.HBox([R10_back_button, R10_excel_standalone_button])
    ], layout={
        'padding': '30px',
        'max_width': '1400px',
        'margin': '0 auto',
        'background': '#f5f5f5'
    })

    R10_back_button.style._css = {'transition': 'background-color 0.3s'}
    R10_back_button.style._hover = {'background_color': '#e0e0e0'}
    R10_excel_standalone_button.style._css = {'transition': 'background-color 0.3s'}
    R10_excel_standalone_button.style._hover = {'background_color': '#1976D2'}

    display(r10_layout)

    # Refresh dashboard on load
    refresh_dashboard()

# Initialize by showing the Login tab
showPage("Login")

# A simple function for jumping between tabs
def showPage(tabName):
  with tb_all.output_to(tabName):
    display()
    # Refresh the page content when navigating to specific tabs
    if tabName == "Manager Dashboard":
        refresh_dashboard()
    elif tabName == "View All Orders":
        refresh_orders_page()
    elif tabName == "Sales Revenue by Date":
        refresh_r5_page()
    elif tabName == "Sales Quantity by Date":
        refresh_r6_page()
    elif tabName == "Trade-in Discounts by Date":
        refresh_r7_page()
    elif tabName == "Old Devices Received":
        refresh_r8_page()
    elif tabName == "Out-of-Stock Products":
        refresh_r9_page()
    elif tabName == "Top 10 Customers":
        refresh_r10_page()

# Initialize by showing the Login tab
showPage("Login")

c.close()
